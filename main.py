# -*- coding: utf-8 -*-
import asyncio
import os
import logging
import pytz 
import io
import aiohttp
import json
import csv
import calendar
import re
import requests  # Shrift yuklash uchun
from datetime import datetime, timedelta, date as d_date, time as d_time
from collections import defaultdict
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart, Command
from aiogram.utils.keyboard import InlineKeyboardBuilder, ReplyKeyboardBuilder
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardButton
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from geopy.distance import geodesic
from aiohttp import web
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import asyncpg
import pickle
import traceback

# --- LOGGING ---
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

# --- SHRIFTNI AVTOMATIK YUKLASH (KOREYS VA RUS UCHUN) ---
FONT_PATH = "NanumGothic.ttf"
if not os.path.exists(FONT_PATH):
    try:
        # Agar shrift fayli bo'lmasa, uni internetdan yuklab olamiz (faqat 1 marta)
        url = "https://github.com/google/fonts/raw/main/ofl/nanumgothic/NanumGothic-Regular.ttf"
        logging.info("⬇️ NanumGothic shrifti yuklanmoqda...")
        r = requests.get(url, timeout=30)
        with open(FONT_PATH, "wb") as f:
            f.write(r.content)
        logging.info("✅ NanumGothic shrifti yuklandi")
    except Exception as e:
        logging.error(f"❌ Shrift yuklashda xatolik: {e}")

try:
    pdfmetrics.registerFont(TTFont('Nanum', FONT_PATH))
    FONT_NAME = 'Nanum'
    FONT_NAME_BOLD = 'Nanum'
    logging.info("✅ Nanum shrifti ro'yxatdan o'tkazildi")
except:
    # Agar shrift topilmasa standartga qaytadi
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        FONT_NAME = 'DejaVu'
        FONT_NAME_BOLD = 'DejaVu-Bold'
        logging.info("✅ DejaVu shrifti yuklandi")
    except:
        FONT_NAME = 'Helvetica'
        FONT_NAME_BOLD = 'Helvetica-Bold'
        logging.warning("⚠️ Maxsus shrift topilmadi, Helvetica ishlatiladi")

# --- EMOJI TOZALASH FUNKSIYASI (PDF UCHUN) ---
def clean_pdf_text(text: str) -> str:
    """PDF shrifti tanimaydigan emojilarni olib tashlaydi, harflarni qoldiradi"""
    if not text: 
        return ""
    # Faqat Lotin, Kirill (Rus) va Koreys (Hangul) harflarini qoldiruvchi regex
    return re.sub(r'[^\w\s\d\.\:\-\/\[\]\(\)\|\!\?\,\u0400-\u04FF\uAC00-\uD7AF\u1100-\u11FF]+', '', text)

# --- SOZLAMALAR ---
TOKEN = os.environ.get("BOT_TOKEN")
if not TOKEN:
    raise ValueError("BOT_TOKEN topilmadi! Render.com da environment variable qo'shing")
DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise ValueError("DATABASE_URL topilmadi! Render.com da environment variable qo'shing")
ADMIN_GROUP_ID = -1003885800610 
UZB_TZ = pytz.timezone('Asia/Tashkent') 

# --- OB-HAVO SOZLAMALARI ---
WEATHER_API_KEY = os.environ.get("WEATHER_API_KEY")
if not WEATHER_API_KEY:
    raise ValueError("WEATHER_API_KEY topilmadi! Render.com da environment variable qo'shing")
WEATHER_API_URL = "http://api.openweathermap.org/data/2.5/weather"

# Bot obyektini yaratish
bot = Bot(token=TOKEN)

# Dispatcher obyektini yaratish
dp = Dispatcher()

# Foydalanuvchi ma'lumotlari (RAM da vaqtinchalik)
user_names = {}
user_specialty = {}
user_status = {}
user_languages = {}
user_ids = set()
daily_attendance_log = set()
attendance_counter = {}
schedules = {}
user_schedules = defaultdict(list)
broadcast_history = []

# RAM da guruhlar va o'quvchilar (tezroq ishlash uchun, lekin asosiy PostgreSQL'da)
groups = {}  # group_id -> group ma'lumotlari
group_students = defaultdict(list)  # group_id -> o'quvchilar ro'yxati

# BARCHA LOKATSIYALAR RO'YXATI
LOCATIONS =[
    {"name": "Kimyo Xalqaro Universiteti", "lat": 41.257490, "lon": 69.220109},
    {"name": "Menejment Universiteti", "lat": 41.270526, "lon": 69.236492},
    {"name": "78-Maktab", "lat": 41.282791, "lon": 69.173290},
    {"name": "126-Maktab", "lat": 41.260249, "lon": 69.153216},
    {"name": "290-Maktab", "lat": 41.234736, "lon": 69.350745},
    {"name": "348-Maktab", "lat": 41.214092, "lon": 69.340152},
    {"name": "347-Maktab", "lat": 41.236833, "lon": 69.372048},
    {"name": "358-Maktab", "lat": 41.240690, "lon": 69.366529},
    {"name": "346-Maktab", "lat": 41.216158, "lon": 69.323902},
    {"name": "293-Maktab", "lat": 41.253573, "lon": 69.377204},
    {"name": "345-Maktab", "lat": 41.220456, "lon": 69.333441},
    {"name": "IM.Gubkin Litseyi", "lat": 41.254183, "lon": 69.382270},
    {"name": "Narxoz universiteti", "lat": 41.308916, "lon": 69.247496},
    {"name": "Narxoz litseyi", "lat": 41.306951, "lon": 69.247667},
    {"name": "Tekstil litseyi", "lat": 41.284784, "lon": 69.249356},
    {"name": "200-Maktab", "lat": 41.263860, "lon": 69.181538},
    {"name": "Selxoz litseyi", "lat": 41.362532, "lon": 69.340768},
    {"name": "294-Maktab", "lat": 41.281633, "lon": 69.289237},
    {"name": "Umnie Deti School", "lat": 41.315790, "lon": 69.209515},
    {"name": "Cambridge School", "lat": 41.342296, "lon": 69.167571}
]

ALLOWED_DISTANCE = 500

# --- OYLIK HISOBOT UCHUN STATE ---
class MonthlyReport(StatesGroup):
    waiting_for_date_range = State()

# --- OYLIK KALKULYATOR UCHUN STATE ---
class SalaryCalc(StatesGroup):
    selecting_specialty = State()
    selecting_teacher = State()
    # Ko'p filialli hisob uchun
    entering_branch_data = State() 
    entering_students = State()
    entering_lessons = State()
    selecting_percentage = State()
    entering_penalty_it_percent = State()  # IT uchun % jarima
    entering_penalty_kr_sum = State()      # Koreys tili uchun so'mda jarima
    entering_payment = State()              # IT uchun jami tushum

# --- VIZUAL JADVAL UCHUN STATE ---
class VisualSchedule(StatesGroup):
    selecting_branch = State()

# --- GURUH YARATISH UCHUN STATE (YANGI) ---
class CreateGroup(StatesGroup):
    selecting_branch = State()
    selecting_type = State()
    selecting_teacher = State()
    selecting_days = State()
    entering_time = State()
    entering_group_name = State()
    adding_student_name = State()
    adding_student_phone = State()
    confirm_student = State()

# --- O'QUVCHILAR DAVOMATI UCHUN STATE (YANGI) ---
class StudentAttendance(StatesGroup):
    selecting_students = State()

# --- POSTGRESQL DATABASE CLASS ---
class Database:
    def __init__(self, url):
        self.url = url
        self.pool = None
    
    async def create_pool(self):
        try:
            self.pool = await asyncpg.create_pool(self.url)
            logging.info("✅ PostgreSQL ga ulandik!")
            return True
        except Exception as e:
            logging.error(f"❌ PostgreSQL ga ulanishda xato: {e}")
            return False
    
    async def init_tables(self):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    user_id BIGINT PRIMARY KEY,
                    full_name TEXT,
                    specialty TEXT,
                    status TEXT DEFAULT 'active',
                    language TEXT DEFAULT 'uz',
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS attendance (
                    id SERIAL PRIMARY KEY,
                    user_id BIGINT REFERENCES users(user_id),
                    branch TEXT,
                    date DATE,
                    time TIME,
                    UNIQUE(user_id, branch, date)
                )
            """)
            
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS schedules (
                    schedule_id TEXT PRIMARY KEY,
                    user_id BIGINT REFERENCES users(user_id),
                    branch TEXT,
                    lesson_type TEXT,
                    days_data JSONB,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS broadcast_history (
                    id SERIAL PRIMARY KEY,
                    message_text TEXT,
                    sent_count INT,
                    failed_count INT,
                    specialty TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            # Guruhlar va o'quvchilar uchun yangi jadvallar
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS groups (
                    id SERIAL PRIMARY KEY,
                    group_name TEXT,
                    branch TEXT,
                    lesson_type TEXT,
                    teacher_id BIGINT REFERENCES users(user_id),
                    days_data JSONB,
                    time_text TEXT,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """)
            
            await conn.execute("""
                CREATE TABLE IF NOT EXISTS group_students (
                    id SERIAL PRIMARY KEY,
                    group_id INT REFERENCES groups(id) ON DELETE CASCADE,
                    student_name TEXT,
                    student_phone TEXT
                )
            """)
            
            logging.info("✅ Jadvallar yaratildi!")
    
    async def save_user(self, user_id, full_name, specialty=None, language='uz'):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO users (user_id, full_name, specialty, language, status)
                VALUES ($1, $2, $3, $4, 'active')
                ON CONFLICT (user_id) 
                DO UPDATE SET 
                    full_name = EXCLUDED.full_name,
                    specialty = COALESCE(EXCLUDED.specialty, users.specialty),
                    language = EXCLUDED.language
            """, user_id, full_name, specialty, language)
    
    async def update_user_status(self, user_id, status):
        async with self.pool.acquire() as conn:
            await conn.execute("UPDATE users SET status = $1 WHERE user_id = $2", status, user_id)
    
    async def get_all_users(self):
        async with self.pool.acquire() as conn:
            return await conn.fetch("SELECT * FROM users")
    
    async def get_user(self, user_id):
        async with self.pool.acquire() as conn:
            return await conn.fetchrow("SELECT * FROM users WHERE user_id = $1", user_id)
    
    async def save_attendance(self, user_id, branch, att_date, att_time):
        try:
            async with self.pool.acquire() as conn:
                from datetime import datetime, time
                date_obj = datetime.strptime(att_date, "%Y-%m-%d").date()
                time_parts = att_time.split(':')
                time_obj = time(int(time_parts[0]), int(time_parts[1]), int(time_parts[2]))
                
                await conn.execute("""
                    INSERT INTO attendance (user_id, branch, date, time)
                    VALUES ($1, $2, $3, $4)
                    ON CONFLICT (user_id, branch, date) DO NOTHING
                """, user_id, branch, date_obj, time_obj)
                logging.info(f"✅ Davomat saqlandi: user={user_id}, branch={branch}")
        except Exception as e:
            logging.error(f"❌ Davomat saqlashda xato: {e}")
    
    async def get_user_attendance(self, user_id):
        async with self.pool.acquire() as conn:
            return await conn.fetch("""
                SELECT * FROM attendance 
                WHERE user_id = $1 
                ORDER BY date DESC, time DESC
            """, user_id)
    
    async def get_attendance_by_date(self, att_date):
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("""
                SELECT a.*, u.full_name, u.specialty 
                FROM attendance a
                JOIN users u ON a.user_id = u.user_id
                WHERE a.date = $1
                ORDER BY a.time
            """, att_date)
            return rows
    
    async def get_all_attendance(self):
        async with self.pool.acquire() as conn:
            return await conn.fetch("SELECT * FROM attendance")
    
    async def save_schedule(self, schedule_id, user_id, branch, lesson_type, days_dict):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO schedules (schedule_id, user_id, branch, lesson_type, days_data)
                VALUES ($1, $2, $3, $4, $5::jsonb)
            """, schedule_id, user_id, branch, lesson_type, json.dumps(days_dict))
    
    async def get_user_schedules(self, user_id):
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM schedules WHERE user_id = $1", user_id)
            result = []
            for row in rows:
                data = dict(row)
                data['days'] = json.loads(data['days_data'])
                result.append(data)
            return result
    
    async def get_all_schedules(self):
        async with self.pool.acquire() as conn:
            rows = await conn.fetch("SELECT * FROM schedules")
            result = []
            for row in rows:
                data = dict(row)
                data['days'] = json.loads(data['days_data'])
                result.append(data)
            return result
    
    async def delete_schedule(self, schedule_id):
        async with self.pool.acquire() as conn:
            await conn.execute("DELETE FROM schedules WHERE schedule_id = $1", schedule_id)
    
    async def update_schedule(self, schedule_id, branch, lesson_type, days_dict):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                UPDATE schedules 
                SET branch = $1, lesson_type = $2, days_data = $3::jsonb
                WHERE schedule_id = $4
            """, branch, lesson_type, json.dumps(days_dict), schedule_id)
    
    async def save_broadcast(self, message_text, sent_count, failed_count, specialty):
        async with self.pool.acquire() as conn:
            await conn.execute("""
                INSERT INTO broadcast_history (message_text, sent_count, failed_count, specialty)
                VALUES ($1, $2, $3, $4)
            """, message_text, sent_count, failed_count, specialty)
    
    async def load_to_ram(self):
        global user_names, user_specialty, user_status, user_languages, user_ids
        global daily_attendance_log, attendance_counter, schedules, user_schedules
        global groups, group_students
        
        users = await self.get_all_users()
        for u in users:
            user_ids.add(u['user_id'])
            user_names[u['user_id']] = u['full_name']
            user_specialty[u['user_id']] = u['specialty']
            user_status[u['user_id']] = u['status']
            user_languages[u['user_id']] = u['language']
        
        attendances = await self.get_all_attendance()
        for r in attendances:
            daily_attendance_log.add((
                r['user_id'],
                r['branch'],
                r['date'].isoformat(),
                r['time'].strftime("%H:%M:%S")
            ))
            month = r['date'].strftime("%Y-%m")
            key = (r['user_id'], r['branch'], month)
            attendance_counter[key] = attendance_counter.get(key, 0) + 1
        
        all_schedules = await self.get_all_schedules()
        for r in all_schedules:
            schedules[r['schedule_id']] = {
                'user_id': r['user_id'],
                'branch': r['branch'],
                'lesson_type': r['lesson_type'],
                'days': r['days']
            }
            user_schedules[r['user_id']].append(r['schedule_id'])
        
        # Guruh va o'quvchilarni yuklash
        async with self.pool.acquire() as conn:
            all_groups = await conn.fetch("SELECT * FROM groups")
            for g in all_groups:
                groups[g['id']] = {
                    'group_name': g['group_name'],
                    'branch': g['branch'],
                    'lesson_type': g['lesson_type'],
                    'teacher_id': g['teacher_id'],
                    'days': json.loads(g['days_data']),
                    'time': g['time_text']
                }
                
                students = await conn.fetch("SELECT * FROM group_students WHERE group_id = $1", g['id'])
                group_students[g['id']] = [{'name': s['student_name'], 'phone': s['student_phone']} for s in students]
        
        logging.info(f"✅ RAM ga yuklandi: {len(user_ids)} foydalanuvchi, {len(daily_attendance_log)} davomat, {len(groups)} guruh")

db = Database(DATABASE_URL)

class Registration(StatesGroup):
    waiting_for_name = State()
    waiting_for_specialty = State()

class AdminAddSchedule(StatesGroup):
    selecting_teacher = State()
    selecting_branch = State()
    selecting_lesson_type = State()
    selecting_weekdays = State()
    entering_time = State()

class AdminEditSchedule(StatesGroup):
    selecting_schedule = State()
    editing_branch = State()
    editing_lesson_type = State()
    editing_weekdays = State()
    editing_time = State()

class Broadcast(StatesGroup):
    selecting_specialty = State()
    waiting_for_message = State()
    waiting_for_confirm = State()

class AddLocation(StatesGroup):
    waiting_for_name = State()
    waiting_for_coords = State()

class PDFReport(StatesGroup):
    waiting_for_date = State()

class ProfileEdit(StatesGroup):
    waiting_for_new_name = State()

class AdminPDFReport(StatesGroup):
    waiting_for_report_type = State()

WEEKDAYS = {
    'uz':['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba', 'Yakshanba'],
    'ru':['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье'],
    'kr':['월요일', '화요일', '수요일', '목요일', '금요일', '토요일', '일요일']
}

WEEKDAY_ORDER = {
    'Dushanba': 0, 'Seshanba': 1, 'Chorshanba': 2, 'Payshanba': 3, 'Juma': 4, 'Shanba': 5, 'Yakshanba': 6
}

LESSON_TYPES = {
    'uz':['IT', 'Koreys tili'],
    'ru': ['IT', 'Корейский язык'],
    'kr': ['IT', '한국어']
}

WEEKDAYS_UZ =["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]

WEATHER_RECOMMENDATIONS = {
    "Clear": {
        "uz": "☀️ Bugun havo ochiq. Sayr qilish uchun ajoyib kun!",
        "ru": "☀️ Сегодня ясно. Отличный день для прогулки!",
        "kr": "☀️ 오늘은 맑은 날씨입니다. 산책하기 좋은 날이에요!"
    },
    "Clouds": {
        "uz": "☁️ Bugun havo bulutli. Salqin havo bilan ish kuningiz samarali o'tsin!",
        "ru": "☁️ Сегодня облачно. Пусть прохладная погода сделает ваш рабочий день продуктивным!",
        "kr": "☁️ 오늘은 흐린 날씨입니다. 시원한 날씨와 함께 즐거운 하루 되세요!"
    },
    "Rain": {
        "uz": "🌧️ Bugun yomg'ir yog'moqda. Soyabon olishni unutmang!",
        "ru": "🌧️ Сегодня идет дождь. Не забудьте взять зонтик!",
        "kr": "🌧️ 오늘은 비가 옵니다. 우산 챙기는 것 잊지마세요!"
    },
    "Thunderstorm": {
        "uz": "⛈️ Momaqaldiroq bo'lmoqda. Ehtiyot bo'ling!",
        "ru": "⛈️ Гроза. Будьте осторожны!",
        "kr": "⛈️ 천둥번개가 칩니다. 조심하세요!"
    },
    "Snow": {
        "uz": "❄️ Qor yog'moqda. Issiq kiyining!",
        "ru": "❄️ Идет снег. Одевайтесь теплее!",
        "kr": "❄️ 눈이 옵니다. 따뜻하게 입으세요!"
    },
    "Mist": {
        "uz": "🌫️ Tuman tushgan. Haydovchilar ehtiyot bo'ling!",
        "ru": "🌫️ Туман. Водители, будьте осторожны!",
        "kr": "🌫️ 안개가 끼었습니다. 운전자분들 조심하세요!"
    },
    "Fog": {
        "uz": "🌫️ Tuman tushgan. Haydovchilar ehtiyot bo'ling!",
        "ru": "🌫️ Туман. Водители, будьте осторожны!",
        "kr": "🌫️ 안개가 끼었습니다. 운전자분들 조심하세요!"
    },
    "Haze": {
        "uz": "🌫️ Havo tumanli. Ehtiyot bo'ling!",
        "ru": "🌫️ Дымка. Будьте осторожны!",
        "kr": "🌫️ 연무가 끼었습니다. 조심하세요!"
    }
}

TRANSLATIONS = {
    'uz': {
        'welcome': "🌟 HANCOM ACADEMYning o'qituvchilar uchun davomat botiga hush kelibsiz, {name}!",
        'ask_name': "👤 Iltimos, ism va familiyangizni kiriting:\n\nMasalan: Ali Karimov",
        'ask_specialty': "📚 Qaysi fan o'qituvchisisiz?",
        'specialty_it': "💻 IT",
        'specialty_korean': "🇰🇷 Koreys tili",
        'specialty_office': "🏢 Ofis xodimi",
        'stats': "📊 Sizning statistikangiz:",
        'no_stats': "💭 Hali davomat qilmagansiz",
        'branches': "🏢 Mavjud filiallar (lokatsiya):",
        'help': "🤖 Botdan foydalanish qo'llanmasi:\n\n📍 Davomat qilish uchun:\n• Pastdagi \"📍 Kelganimni tasdiqlash\" tugmasini bosing\n• Joylashuvingizni yuboring\n\n📊 Statistika:\n• \"📊 Mening statistikam\" - shaxsiy davomat tarixingiz\n• \"🏢 Filiallar\" - barcha mavjud filiallar ro'yxati\n\n⚠️ Eslatmalar:\n• Har bir filialda kuniga faqat 1 marta davomat qilish mumkin\n• Davomat faqat Toshkent vaqti bilan hisoblanadi",
        'attendance_success': "✅ Davomat tasdiqlandi!\n\n🏫 Filial: {branch}\n📅 Sana: {date}\n⏰ Vaqt: {time}\n📊 Bu oydagi tashriflar: {count} marta\n📍 Masofa: {distance:.1f} metr",
        'already_attended': "⚠️ Siz bugun {branch} hududida allaqachon davomatdan o'tgansiz!",
        'not_in_area': "❌ Siz belgilangan ta'lim muassasalari hududida emassiz!",
        'daily_reminder': "⏰ Eslatma! Bugun hali davomat qilmagansiz. Ish kuningizni boshlash uchun davomatni tasdiqlang!",
        'weekly_top': "🏆 Haftaning eng faol o'qituvchilari:\n\n{top_list}",
        'monthly_report': "📊 {month} oyi uchun hisobot\n\n{report}",
        'language_changed': "✅ Til o'zgartirildi: O'zbek tili",
        'language_prompt': "Iltimos, tilni tanlang:",
        'view_schedules': "📋 Dars jadvalim (PDF)",
        'my_schedule': "📅 Sizning dars jadvalingiz PDF formatida tayyorlandi!",
        'no_schedules': "📭 Sizga hali dars jadvali biriktirilmagan.",
        'schedule_updated': "📢 Sizning dars jadvalingiz yangilandi!",
        'schedule_deleted_notify': "📢 Sizning dars jadvalingiz o'chirildi.",
        'reminder': "⏰ Eslatma!\n\nBugun soat {time} da {branch} filialida darsingiz bor.\nDavomat qilishni unutmang!",
        'lesson_started_attended': "✅ Dars boshlandi va siz muvaffaqiyatli davomatni amalga oshirdingiz!\n\nE'tiboringizni darsga qaratishingiz mumkin.\nDarsga kelgan o'quvchilarni davomat qilishni yodingizdan chiqarmang.\n\nHayrli kun!",
        'lesson_started_not_attended': "⚠️ Sizning darsingiz boshlandi, lekin hali davomat qilmadingiz!\n\n📌 {branch} filialida soat {time} da darsingiz boshlangan.\n📍 Iltimos, darhol davomat qiling yoki sababini admin xabardor qiling.\n\nDavomat qilish uchun 📍 Kelganimni tasdiqlash tugmasini bosing.",
        'select_teacher': "👤 O'qituvchini tanlang:",
        'select_lesson_type': "📚 Dars turini tanlang:",
        'active_schedules': "📋 Faol dars jadvallari",
        'no_active_schedules': "📭 Hali dars jadvallari mavjud emas.",
        'schedule_info': "{teacher} [{specialty}]\n🏢 {branch}\n📚 {lesson_type}\n{days_times}",
        'enter_date': "📅 Hisobot olish uchun sanani kiriting (format: YYYY-MM-DD)\nMasalan: 2026-03-01",
        'invalid_date': "❌ Noto'g'ri sana formati. Qaytadan urinib ko'ring:",
        'select_broadcast_specialty': "📢 Qaysi fan o'qituvchilariga xabar yubormoqchisiz?",
        'all_teachers': "👥 Hammasi",
        'edit_schedule': "✏️ Dars jadvalini tahrirlash",
        'select_new_branch': "🏢 Yangi filialni tanlang:",
        'select_new_lesson_type': "📚 Yangi dars turini tanlang:",
        'select_new_weekdays': "📅 Yangi kunlarni tanlang:",
        'enter_new_time': "⏰ {weekday} kuni uchun yangi vaqtni kiriting:\n\nFormat: HH:MM (masalan: 09:00)",
        'ontime': "Vaqtida",
        'late': "Kechikkan",
        'my_profile': "👤 Mening profilim",
        'profile_info': "👤 Sizning profilingiz:\n\nIsm: {name}\nMutaxassislik: {specialty}\nTil: {lang}",
        'edit_name': "✏️ Ismni o'zgartirish",
        'edit_my_specialty': "📚 Faoliyat turini o'zgartirish",
        'enter_new_name': "Yangi ism va familiyangizni kiriting:",
        'name_updated': "✅ Ismingiz muvaffaqiyatli yangilandi!",
        'back_to_menu': "🔙 Menyuga qaytish",
        'select_new_spec': "Yangi faoliyat turini tanlang:",
        'spec_updated': "✅ Mutaxassislik yangilandi!",
        'back_btn': "🔙 Ortga",
        'pdf_title': "Dars Jadvali",
        'pdf_headers': ['Kun', 'Vaqt'],
        'pdf_created': "Yaratilgan sana",
        'group_students_title': "O'quvchilar davomati",
        'group_students_submit': "Davomatni yuborish",
        'group_students_sent': "✅ O'quvchilar davomati yuborildi!",
        'buttons': {
            'attendance': "📍 Kelganimni tasdiqlash",
            'my_stats': "📊 Mening statistikam",
            'branches': "🏢 Filiallar",
            'top_week': "🏆 Hafta topi",
            'view_schedules': "📋 Dars jadvalim (PDF)",
            'help': "❓ Yordam",
            'language': "🌐 Til"
        }
    },
    'ru': {
        'welcome': "🌟 Добро пожаловать в бот для отметок HANCOM ACADEMY для учителей, {name}!",
        'ask_name': "👤 Пожалуйста, введите ваше имя и фамилию:\n\nНапример: Ali Karimov",
        'ask_specialty': "📚 Какой предмет вы преподаете?",
        'specialty_it': "💻 IT",
        'specialty_korean': "🇰🇷 Корейский язык",
        'specialty_office': "🏢 Офисный сотрудник",
        'stats': "📊 Ваша статистика:",
        'no_stats': "💭 Вы еще не отмечались",
        'branches': "🏢 Доступные филиалы (локация):",
        'help': "🤖 Руководство по использования:\n\n📍 Для отметки:\n• Нажмите кнопку \"📍 Подтвердить прибытие\"\n• Отправьте свою геолокацию\n\n📊 Статистика:\n• \"📊 Моя статистика\" - история отметок\n• \"🏢 Филиалы\" - список всех филиалов\n\n⚠️ Примечания:\n• В каждом филиале можно отмечаться только 1 раз в день\n• Отметки записываются по ташкентскому времени",
        'attendance_success': "✅ Отметка подтверждена!\n\n🏫 Филиал: {branch}\n📅 Дата: {date}\n⏰ Время: {time}\n📊 Посещений в этом месяце: {count}\n📍 Расстояние: {distance:.1f} м",
        'already_attended': "⚠️ Вы уже отмечались сегодня в филиале {branch}!",
        'not_in_area': "❌ Вы не находитесь в зоне учебных заведений!",
        'daily_reminder': "⏰ Напоминание! Вы еще не отметились сегодня. Подтвердите свое прибытие для начала рабочего дня!",
        'weekly_top': "🏆 Самые активные учителя недели:\n\n{top_list}",
        'monthly_report': "📊 Отчет за {month}\n\n{report}",
        'language_changed': "✅ Язык изменен: Русский язык",
        'language_prompt': "Пожалуйста, выберите язык:",
        'view_schedules': "📋 Мое расписание (PDF)",
        'my_schedule': "📅 Ваше расписание уроков готово в формате PDF!",
        'no_schedules': "📭 Вам еще не назначено расписание.",
        'schedule_updated': "📢 Ваше расписание обновлено!",
        'schedule_deleted_notify': "📢 Ваше расписание удалено.",
        'reminder': "⏰ Напоминание!\n\nСегодня в {time} у вас урок в филиале {branch}.\nНе забудьте отметиться!",
        'lesson_started_attended': "✅ Урок начался и вы успешно отметились!\n\nМожете сосредоточиться на уроке.\nНе забудьте отметить присутствующих учеников.\n\nХорошего дня!",
        'lesson_started_not_attended': "⚠️ Ваш урок начался, но вы еще не отметились!\n\n📌 В филиале {branch} в {time} начался ваш урок.\n📍 Пожалуйста, немедленно отметьтесь или сообщите причину администратору.\n\nДля отметки нажмите кнопку 📍 Подтвердить прибытие.",
        'select_teacher': "👤 Выберите учителя:",
        'select_lesson_type': "📚 Выберите тип урока:",
        'active_schedules': "📋 Активные расписания",
        'no_active_schedules': "📭 Нет активных расписаний.",
        'schedule_info': "{teacher}[{specialty}]\n🏢 {branch}\n📚 {lesson_type}\n{days_times}",
        'enter_date': "📅 Введите дату для отчета (формат: YYYY-MM-DD)\nНапример: 2026-03-01",
        'invalid_date': "❌ Неверный формат даты. Попробуйте снова:",
        'select_broadcast_specialty': "📢 Каким учителям отправить сообщение?",
        'all_teachers': "👥 Всем",
        'edit_schedule': "✏️ Редактирование расписания",
        'select_new_branch': "🏢 Выберите новый филиал:",
        'select_new_lesson_type': "📚 Выберите новый тип урока:",
        'select_new_weekdays': "📅 Выберите новые дни:",
        'enter_new_time': "⏰ Введите новое время для {weekday}:\n\nФормат: HH:MM (например: 09:00)",
        'ontime': "Вовремя",
        'late': "Опоздал",
        'my_profile': "👤 Мой профиль",
        'profile_info': "👤 Ваш профиль:\n\nИмя: {name}\nСпециальность: {specialty}\nЯзык: {lang}",
        'edit_name': "✏️ Изменить имя",
        'edit_my_specialty': "📚 Изменить направление",
        'enter_new_name': "Введите новое имя и фамилию:",
        'name_updated': "✅ Ваше имя успешно обновлено!",
        'back_to_menu': "🔙 Вернуться в меню",
        'select_new_spec': "Выберите новое направление:",
        'spec_updated': "✅ Специальность обновлена!",
        'back_btn': "🔙 Назад",
        'pdf_title': "Расписание занятий",
        'pdf_headers': ['День', 'Время'],
        'pdf_created': "Дата создания",
        'group_students_title': "Отметка учеников",
        'group_students_submit': "Отправить отметку",
        'group_students_sent': "✅ Отметка учеников отправлена!",
        'buttons': {
            'attendance': "📍 Подтвердить прибытие",
            'my_stats': "📊 Моя статистика",
            'branches': "🏢 Филиалы",
            'top_week': "🏆 Топ недели",
            'view_schedules': "📋 Мое расписание (PDF)",
            'help': "❓ Помощь",
            'language': "🌐 Язык"
        }
    },
    'kr': {
        'welcome': "🌟 HANCOM ACADEMY 교사용 출석 체크 봇에 오신 것을 환영합니다, {name}!",
        'ask_name': "👤 이름과 성을 입력하세요:\n\n예: Ali Karimov",
        'ask_specialty': "📚 어떤 과목을 가르치시나요?",
        'specialty_it': "💻 IT",
        'specialty_korean': "🇰🇷 한국어",
        'specialty_office': "🏢 사무원",
        'stats': "📊 내 통계:",
        'no_stats': "💭 아직 출석 체크하지 않았습니다",
        'branches': "🏢 등록된 지점 (위치):",
        'help': "🤖 사용 설명서:\n\n📍 출석 체크 방법:\n• 하단의 \"📍 출석 확인\" 버튼을 누르세요\n• 위치를 전송하세요\n\n📊 통계:\n• \"📊 내 통계\" - 개인 출석 기록\n• \"🏢 지점\" - 모든 지점 목록\n\n⚠️ 참고사항:\n• 각 지점에서 하루에 한 번만 출석 체크 가능\n• 출석은 타슈켄트 시간 기준으로 기록됨",
        'attendance_success': "✅ 출석이 확인되었습니다!\n\n🏫 지점: {branch}\n📅 날짜: {date}\n⏰ 시간: {time}\n📊 이번 달 출석: {count}회\n📍 거리: {distance:.1f}미터",
        'already_attended': "⚠️ 오늘 이미 {branch} 지점에서 출석 체크하셨습니다!",
        'not_in_area': "❌ 지정된 교육 기관 구역 내에 있지 않습니다!",
        'daily_reminder': "⏰ 알림! 오늘 아직 출석 체크하지 않으셨습니다. 업무 시작을 위해 출석을 확인하세요!",
        'weekly_top': "🏆 이번 주 가장 활발한 교사:\n\n{top_list}",
        'monthly_report': "📊 {month}월 보고서\n\n{report}",
        'language_changed': "✅ 언어가 변경되었습니다: 한국어",
        'language_prompt': "언어를 선택하세요:",
        'view_schedules': "📋 내 시간표 (PDF)",
        'my_schedule': "📅 내 수업 시간표가 PDF 형식으로 준비되었습니다!",
        'no_schedules': "📭 아직 시간표가 배정되지 않았습니다.",
        'schedule_updated': "📢 시간표가 업데이트되었습니다!",
        'schedule_deleted_notify': "📢 시간표가 삭제되었습니다.",
        'reminder': "⏰ 알림!\n\n오늘 {time}에 {branch} 지점에서 수업이 있습니다.\n출석 체크를 잊지 마세요!",
        'lesson_started_attended': "✅ 수업이 시작되었고 출석이 확인되었습니다!\n\n수업에 집중하세요.\n학생들 출석 체크하는 것을 잊지 마세요.\n\n좋은 하루 되세요!",
        'lesson_started_not_attended': "⚠️ 수업이 시작되었지만 아직 출석 체크하지 않으셨습니다!\n\n📌 {branch} 지점에서 {time}에 수업이 시작되었습니다.\n📍 즉시 출석 체크하거나 관리자에게 사유를 알려주세요.\n\n출석 체크를 위해 📍 출석 확인 버튼을 누르세요.",
        'select_teacher': "👤 교사를 선택하세요:",
        'select_lesson_type': "📚 수업 유형을 선택하세요:",
        'active_schedules': "📋 활성 시간표",
        'no_active_schedules': "📭 활성 시간표가 없습니다.",
        'schedule_info': "{teacher} [{specialty}]\n🏢 {branch}\n📚 {lesson_type}\n{days_times}",
        'enter_date': "📅 보고서 날짜를 입력하세요 (형식: YYYY-MM-DD)\n예: 2026-03-01",
        'invalid_date': "❌ 잘못된 날짜 형식입니다. 다시 시도하세요:",
        'select_broadcast_specialty': "📢 어떤 선생님들에게 메시지를 보낼까요?",
        'all_teachers': "👥 모두",
        'edit_schedule': "✏️ 시간표 편집",
        'select_new_branch': "🏢 새 지점을 선택하세요:",
        'select_new_lesson_type': "📚 새 수업 유형을 선택하세요:",
        'select_new_weekdays': "📅 새 요일을 선택하세요:",
        'enter_new_time': "⏰ {weekday} 요일의 새 시간을 입력하세요:\n\n형식: HH:MM (예: 09:00)",
        'ontime': "정시",
        'late': "지각",
        'my_profile': "👤 내 프로필",
        'profile_info': "👤 내 프로필:\n\n이름: {name}\n전공: {specialty}\n언어: {lang}",
        'edit_name': "✏️ 이름 변경",
        'edit_my_specialty': "📚 전공 변경",
        'enter_new_name': "새 이름과 성을 입력하세요:",
        'name_updated': "✅ 이름이 업데이트되었습니다!",
        'back_to_menu': "🔙 메뉴로 돌아가기",
        'select_new_spec': "새 전공을 선택하세요:",
        'spec_updated': "✅ 전공이 업데이트되었습니다!",
        'back_btn': "🔙 뒤로 가기",
        'pdf_title': "수업 시간표",
        'pdf_headers': ["요일", "시간"],
        'pdf_created': "작성일",
        'group_students_title': "학생 출석 체크",
        'group_students_submit': "출석 보내기",
        'group_students_sent': "✅ 학생 출석이 전송되었습니다!",
        'buttons': {
            'attendance': "📍 출석 확인",
            'my_stats': "📊 내 통계",
            'branches': "🏢 지점",
            'top_week': "🏆 주간 TOP",
            'view_schedules': "📋 내 시간표 (PDF)",
            'help': "❓ 도움말",
            'language': "🌐 언어"
        }
    }
}

def get_text(user_id: int, key: str, **kwargs):
    lang = user_languages.get(user_id, 'uz')
    text = TRANSLATIONS[lang].get(key, TRANSLATIONS['uz'].get(key, ''))
    if kwargs:
        try:
            text = text.format(**kwargs)
        except:
            pass
    return text

def get_button_text(user_id: int, button_key: str):
    lang = user_languages.get(user_id, 'uz')
    return TRANSLATIONS[lang]['buttons'][button_key]

def check_admin(chat_id):
    return chat_id == ADMIN_GROUP_ID

def get_specialty_display(specialty: str, lang: str = 'uz') -> str:
    if specialty == 'IT':
        if lang == 'uz':
            return "💻 IT"
        elif lang == 'ru':
            return "💻 IT"
        else:
            return "💻 IT"
    elif specialty == 'Koreys tili':
        if lang == 'uz':
            return "🇰🇷 Koreys tili"
        elif lang == 'ru':
            return "🇰🇷 Корейский язык"
        else:
            return "🇰🇷 한국어"
    else:
        if lang == 'uz':
            return "🏢 Ofis xodimi"
        elif lang == 'ru':
            return "🏢 Офисный сотрудник"
        else:
            return "🏢 사무원"

def sort_weekdays(days_dict):
    order = {'Dushanba': 0, 'Seshanba': 1, 'Chorshanba': 2, 'Payshanba': 3, 'Juma': 4, 'Shanba': 5, 'Yakshanba': 6}
    return dict(sorted(days_dict.items(), key=lambda x: order.get(x[0], 7)))

def calculate_lateness(attendance_time: str, lesson_time: str) -> tuple:
    try:
        att_parts = list(map(int, attendance_time.split(':')))
        les_parts = list(map(int, lesson_time.split(':')))
        
        att_seconds = att_parts[0] * 3600 + att_parts[1] * 60 + att_parts[2]
        les_seconds = les_parts[0] * 3600 + les_parts[1] * 60
        
        diff = att_seconds - les_seconds
        
        if diff <= 60:
            return True, 0
        else:
            return False, int(diff / 60)
    except Exception as e:
        logging.error(f"calculate_lateness error: {e}")
        return True, 0

def get_kr_exam_penalty(perc: int) -> int:
    """Koreys tili imtixon foiziga qarab jarima qaytaradi"""
    if perc < 10: return 900000
    elif perc < 20: return 800000
    elif perc < 30: return 700000
    elif perc < 40: return 600000
    elif perc < 50: return 500000
    elif perc < 60: return 400000
    elif perc < 70: return 300000
    elif perc < 80: return 200000
    elif perc < 90: return 100000
    else: return 0

async def get_combined_report_pdf(report_date: d_date) -> io.BytesIO:
    report_date_str = report_date.strftime("%Y-%m-%d")
    report_weekday = WEEKDAYS_UZ[report_date.weekday()]
    
    check_ins = [list(att) for att in daily_attendance_log if att[2] == report_date_str]

    for s_id, s_data in schedules.items():
        uid = s_data['user_id']
        branch = s_data['branch']
        if report_weekday in s_data['days']:
            already_noted = any(c[0] == uid and c[1] == branch for c in check_ins)
            if not already_noted:
                check_ins.append([uid, branch, report_date_str, "00:00:00", "ABSENT"])

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), topMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=1,
        spaceAfter=20,
        fontName=FONT_NAME_BOLD
    )
    title = Paragraph(f"Davomat Hisoboti: {report_date.strftime('%d.%m.%Y')} ({report_weekday})", title_style)
    elements.append(title)
    elements.append(Spacer(1, 15))

    data = [['№', 'Davomat', 'Dars', 'O\'qituvchi', 'Mutaxassislik', 'Filial', 'Holat', 'Kechikish']]
    
    for i, att in enumerate(sorted(check_ins, key=lambda x: x[3] if x[3] != "00:00:00" else "23:59:59"), 1):
        uid = att[0]
        branch = att[1]
        att_time = att[3]
        is_absent = len(att) > 4 and att[4] == "ABSENT"
        
        teacher_name = user_names.get(uid, "Noma'lum")
        specialty = user_specialty.get(uid, "")
        
        lesson_time = "—"
        for s_id in user_schedules.get(uid, []):
            s = schedules.get(s_id)
            if s and s['branch'] == branch and report_weekday in s['days']:
                lesson_time = s['days'][report_weekday]
                break
        
        if is_absent or att_time == "00:00:00":
            status = "KELMAGAN"
            late_text = "—"
            att_time_disp = "—"
        else:
            ontime, mins = calculate_lateness(att_time, lesson_time)
            status = "Vaqtida" if ontime else "Kechikkan"
            late_text = "0" if ontime else f"{mins} min"
            att_time_disp = att_time

        data.append([str(i), att_time_disp, lesson_time, teacher_name, specialty, branch, status, late_text])

    table = Table(data, colWidths=[0.4*inch, 0.9*inch, 0.9*inch, 1.8*inch, 1.0*inch, 1.5*inch, 1.0*inch, 0.8*inch])

    style_commands = [
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2E86AB')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,-1), FONT_NAME),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]

    for row_idx, row_data in enumerate(data[1:], start=1):
        status = row_data[6]
        if status == "Kechikkan":
            style_commands.append(('BACKGROUND', (6, row_idx), (6, row_idx), colors.HexColor('#FFCDD2')))
            style_commands.append(('TEXTCOLOR', (6, row_idx), (6, row_idx), colors.red))
        elif status == "Vaqtida":
            style_commands.append(('TEXTCOLOR', (6, row_idx), (6, row_idx), colors.green))
        elif status == "KELMAGAN":
            style_commands.append(('BACKGROUND', (6, row_idx), (6, row_idx), colors.HexColor('#EF5350')))
            style_commands.append(('TEXTCOLOR', (6, row_idx), (6, row_idx), colors.white))

    table.setStyle(TableStyle(style_commands))
    
    elements.append(table)
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

async def main_keyboard(user_id: int):
    builder = ReplyKeyboardBuilder()
    builder.add(
        KeyboardButton(text=get_button_text(user_id, 'attendance'), request_location=True),
        KeyboardButton(text=get_button_text(user_id, 'my_stats')),
        KeyboardButton(text=get_button_text(user_id, 'branches')),
        KeyboardButton(text=get_button_text(user_id, 'top_week')),
        KeyboardButton(text=get_button_text(user_id, 'view_schedules')),
        KeyboardButton(text=get_text(user_id, 'my_profile')),
        KeyboardButton(text=get_button_text(user_id, 'help')),
        KeyboardButton(text=get_button_text(user_id, 'language'))
    )
    builder.adjust(1, 2, 2, 2, 2)
    return builder.as_markup(resize_keyboard=True)

async def language_selection_keyboard():
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="🇺🇿 O'zbekcha", callback_data="lang_uz"),
        InlineKeyboardButton(text="🇷🇺 Русский", callback_data="lang_ru"),
        InlineKeyboardButton(text="🇰🇷 한국어", callback_data="lang_kr")
    )
    return builder.as_markup()

async def specialty_keyboard(user_id: int):
    lang = user_languages.get(user_id, 'uz')
    builder = ReplyKeyboardBuilder()
    builder.add(
        KeyboardButton(text=TRANSLATIONS[lang]['specialty_it']),
        KeyboardButton(text=TRANSLATIONS[lang]['specialty_korean']),
        KeyboardButton(text=TRANSLATIONS[lang]['specialty_office'])
    )
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)

def get_yandex_maps_link(lat: float, lon: float) -> str:
    return f"https://yandex.com/maps/?pt={lon},{lat}&z=17&l=map"

# --- STANDARTLASHTIRILGAN PDF FUNKSIYASI (XUDDI ADMIN PANELIDAGIDEK) ---
async def create_schedule_pdf(user_id: int) -> io.BytesIO:
    pdf_buffer = io.BytesIO()
    # Landscape format (A4 yotqizilgan) va keng hoshiyalar
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=landscape(A4), 
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30
    )
    elements = []
    styles = getSampleStyleSheet()
    lang = user_languages.get(user_id, 'uz')
    
    # Shriftni majburiy belgilash
    title_font = FONT_NAME_BOLD
    
    # Sarlavha uslubi (Asosiy sarlavha ko'k rangda)
    title_style = ParagraphStyle(
        'T', fontName=title_font, fontSize=22, alignment=1, 
        spaceAfter=20, textColor=colors.HexColor('#1A237E')
    )
    
    # Oddiy matn uslubi
    normal_style = ParagraphStyle('N', fontName=FONT_NAME, fontSize=14, spaceAfter=10)

    # Ma'lumotlarni emojilardan tozalash
    name = clean_pdf_text(user_names.get(user_id, "User"))
    specialty = clean_pdf_text(get_specialty_display(user_specialty.get(user_id, ''), lang))
    
    # 1. ASOSIY SARLAVHA: ISM | SOHA | DARS JADVALI
    title_text = f"<b>{name.upper()}</b> | {specialty} | {clean_pdf_text(TRANSLATIONS[lang]['pdf_title'])}"
    elements.append(Paragraph(title_text, title_style))
    
    sched_ids = user_schedules.get(user_id, [])
    if not sched_ids:
        elements.append(Paragraph(clean_pdf_text(TRANSLATIONS[lang]['no_schedules']), normal_style))
    else:
        for s_id in sched_ids:
            s = schedules.get(s_id)
            if not s: continue
            
            branch = clean_pdf_text(s['branch'])
            l_type = clean_pdf_text(s.get('lesson_type', 'Dars'))
            
            # 2. FILIAL SARLAVHASI (KO'K BANNER)
            branch_p = Paragraph(f"Filial: {branch} ({l_type})", 
                                 ParagraphStyle('B', fontName=title_font, fontSize=16, textColor=colors.white))
            
            br_table = Table([[branch_p]], colWidths=[9.5*inch])
            br_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#283593')),
                ('LEFTPADDING', (0,0), (-1,-1), 20),
                ('TOPPADDING', (0,0), (-1,-1), 10),
                ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ]))
            elements.append(br_table)
            elements.append(Spacer(1, 5))
            
            # 3. JADVAL QISMI - TO'G'RILANGAN VERSIYA
            headers = [clean_pdf_text(h) for h in TRANSLATIONS[lang]['pdf_headers']]
            data = [headers]
            
            # Hafta kunlari (O'zbekcha kalitlar bilan solishtirish)
            uz_days = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
            
            for i, d_uz in enumerate(uz_days):
                if d_uz in s['days']:
                    # Foydalanuvchi tanlagan tildagi kun nomini olamiz
                    day_name_in_lang = WEEKDAYS[lang][i]
                    data.append([clean_pdf_text(day_name_in_lang), str(s['days'][d_uz])])

            if len(data) > 1:
                # Jadval dizayni (Chiroyli och ko'k fon va zebra style)
                t = Table(data, colWidths=[4.75*inch, 4.75*inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8EAF6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1A237E')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, -1), FONT_NAME),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('FONTSIZE', (0, 0), (-1, -1), 13),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                    ('TOPPADDING', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 20))

    # 4. FOOTER: YARATILGAN SANA
    footer_text = f"{clean_pdf_text(TRANSLATIONS[lang]['pdf_created'])}: {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')}"
    elements.append(Paragraph(footer_text, ParagraphStyle('F', fontName=FONT_NAME, fontSize=10, alignment=2, textColor=colors.grey)))
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

# --- O'QUVCHILAR DAVOMATI UCHUN YORDAMCHI FUNKSIYALAR ---
async def get_student_attendance_kb(group_id, selected_indices):
    """Guruh o'quvchilarini tanlash uchun klaviatura yaratish"""
    builder = InlineKeyboardBuilder()
    students = group_students.get(group_id, [])
    lang = 'uz'  # Default til
    
    for i, std in enumerate(students):
        status = "✅ " if i in selected_indices else "⬜ "
        builder.row(InlineKeyboardButton(
            text=f"{status}{std['name']}", 
            callback_data=f"std_check_{i}"
        ))
    
    builder.row(InlineKeyboardButton(
        text=TRANSLATIONS[lang]['group_students_submit'], 
        callback_data="std_submit"
    ))
    return builder.as_markup()

async def handle(request):
    now_uzb = datetime.now(UZB_TZ)
    return web.Response(
        text=f"Bot is running! ✅\n\n"
             f"📅 Sana: {now_uzb.strftime('%Y-%m-%d')}\n"
             f"⏰ Vaqt: {now_uzb.strftime('%H:%M:%S')}\n"
             f"👥 Foydalanuvchilar: {len(user_ids)} ta\n"
             f"📊 Bugungi davomatlar: {len([k for k in daily_attendance_log if k[2] == now_uzb.strftime('%Y-%m-%d')])} ta\n"
             f"👥 Guruhlar: {len(groups)} ta"
    )

async def health_check(request):
    now_uzb = datetime.now(UZB_TZ)
    logging.info(f"Cron-job.org tomonidan tekshirildi: {now_uzb.strftime('%Y-%m-%d %H:%M:%S')}")
    return web.Response(text=f"Bot healthy - {now_uzb.strftime('%H:%M:%S')}", status=200)

async def start_web_server():
    app = web.Application()
    app.router.add_get('/', handle)
    app.router.add_get('/health', health_check)
    runner = web.AppRunner(app)
    await runner.setup()
    port = int(os.environ.get("PORT", 10000))
    site = web.TCPSite(runner, '0.0.0.0', port)
    await site.start()
    logging.info(f"Web server started on port {port}")

@dp.message(CommandStart())
async def cmd_start(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    if user_id not in user_names:
        if user_id not in user_languages:
            keyboard = await language_selection_keyboard()
            await message.answer(
                "Iltimos, tilni tanlang:\nПожалуйста, выберите язык:\n언어를 선택하세요:",
                reply_markup=keyboard
            )
            return
        
        await state.set_state(Registration.waiting_for_name)
        await message.answer(get_text(user_id, 'ask_name'))
        return
    
    user_ids.add(user_id)
    keyboard = await main_keyboard(user_id)
    name = user_names.get(user_id, message.from_user.full_name)
    specialty = user_specialty.get(user_id, '')
    
    welcome_text = get_text(user_id, 'welcome', name=name)
    if specialty:
        specialty_display = get_specialty_display(specialty, user_languages.get(user_id, 'uz'))
        welcome_text += f"\n\n{specialty_display}"
    
    await message.answer(
        welcome_text,
        reply_markup=keyboard
    )

@dp.message(Registration.waiting_for_name)
async def process_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    full_name = message.text.strip()
    
    user_names[user_id] = full_name
    user_ids.add(user_id)
    
    lang = user_languages.get(user_id, 'uz')
    await db.save_user(user_id, full_name, None, lang)
    
    await state.update_data(name=full_name)
    
    keyboard = await specialty_keyboard(user_id)
    await state.set_state(Registration.waiting_for_specialty)
    await message.answer(
        get_text(user_id, 'ask_specialty'),
        reply_markup=keyboard
    )

@dp.message(Registration.waiting_for_specialty)
async def process_specialty(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    text = message.text
    lang = user_languages.get(user_id, 'uz')
    
    if text == TRANSLATIONS[lang]['specialty_it']:
        specialty = 'IT'
    elif text == TRANSLATIONS[lang]['specialty_korean']:
        specialty = 'Koreys tili'
    elif text == TRANSLATIONS[lang]['specialty_office']:
        specialty = 'Ofis xodimi'
    else:
        await message.answer("❌ Noto'g'ri tanlov. Qaytadan urinib ko'ring.")
        return
    
    user_specialty[user_id] = specialty
    user_status[user_id] = 'active'
    
    await db.save_user(user_id, user_names[user_id], specialty, lang)
    
    await state.clear()
    
    keyboard = await main_keyboard(user_id)
    name = user_names.get(user_id)
    specialty_display = get_specialty_display(specialty, lang)
    
    welcome_text = get_text(user_id, 'welcome', name=name) + f"\n\n{specialty_display}"
    
    await message.answer(
        welcome_text,
        reply_markup=keyboard
    )

@dp.callback_query(F.data.startswith("lang_"))
async def set_initial_language(callback: types.CallbackQuery, state: FSMContext):
    try:
        user_id = callback.from_user.id
        lang = callback.data.split("_")[1]
        
        user_languages[user_id] = lang
        
        await callback.answer()
        await callback.message.delete()
        
        await state.set_state(Registration.waiting_for_name)
        await callback.message.answer(get_text(user_id, 'ask_name'))
    except Exception as e:
        logging.error(f"set_initial_language error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(F.text.in_({'\U0001F310 Til', '\U0001F310 Язык', '\U0001F310 언어'}))
async def change_language(message: types.Message):
    user_id = message.from_user.id
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="🇺🇿 O'zbekcha", callback_data="change_lang_uz"),
        InlineKeyboardButton(text="🇷🇺 Русский", callback_data="change_lang_ru"),
        InlineKeyboardButton(text="🇰🇷 한국어", callback_data="change_lang_kr")
    )
    await message.answer("Tilni tanlang / Выберите язык / 언어를 선택하세요:", reply_markup=builder.as_markup())

@dp.callback_query(F.data.startswith("change_lang_"))
async def set_changed_language(callback: types.CallbackQuery):
    try:
        user_id = callback.from_user.id
        lang = callback.data.split("_")[2]
        user_languages[user_id] = lang
        
        user = await db.get_user(user_id)
        if user:
            await db.save_user(user_id, user['full_name'], user['specialty'], lang)
        
        await callback.answer()
        await callback.message.delete()
        
        keyboard = await main_keyboard(user_id)
        await callback.message.answer(
            get_text(user_id, 'language_changed'),
            reply_markup=keyboard
        )
    except Exception as e:
        logging.error(f"set_changed_language error: {e}")
        await callback.answer("Xatolik yuz berdi")

PROFILE_BTNS = ["👤 Mening profilim", "👤 Мой профиль", "👤 내 프로필"]

@dp.message(F.text.in_(PROFILE_BTNS))
async def show_profile(message: types.Message):
    user_id = message.from_user.id
    lang = user_languages.get(user_id, 'uz')
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    name = user_names.get(user_id, "Noma'lum")
    specialty = user_specialty.get(user_id, "Ko'rsatilmagan")
    
    spec_display = get_specialty_display(specialty, lang)
    lang_display = {'uz': "O'zbekcha", 'ru': "Русский", 'kr': "한국어"}.get(lang, lang)
    
    profile_text = get_text(user_id, 'profile_info', 
                           name=name, 
                           specialty=spec_display, 
                           lang=lang_display)
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text=get_text(user_id, 'edit_name'), callback_data="edit_name"))
    builder.row(InlineKeyboardButton(text=get_text(user_id, 'edit_my_specialty'), callback_data="edit_my_specialty"))
    builder.row(InlineKeyboardButton(text=get_text(user_id, 'back_to_menu'), callback_data="back_to_main"))
    
    await message.answer(
        profile_text,
        reply_markup=builder.as_markup()
    )

@dp.callback_query(F.data == "edit_my_specialty")
async def edit_my_specialty_start(callback: types.CallbackQuery):
    uid = callback.from_user.id
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="💻 IT", callback_data="save_spec_IT"))
    builder.row(InlineKeyboardButton(text="🇰🇷 Koreys tili", callback_data="save_spec_Koreys tili"))
    builder.row(InlineKeyboardButton(text="🏢 Ofis xodimi", callback_data="save_spec_Ofis xodimi"))
    builder.row(InlineKeyboardButton(text=get_text(uid, 'back_btn'), callback_data="back_to_profile_view"))
    
    await callback.message.edit_text(
        get_text(uid, 'select_new_spec'),
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("save_spec_"))
async def save_new_specialty(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    new_spec = callback.data.replace("save_spec_", "")
    
    user_specialty[user_id] = new_spec
    
    try:
        user = await db.get_user(user_id)
        if user:
            await db.save_user(
                user_id=user_id,
                full_name=user['full_name'],
                specialty=new_spec,
                language=user['language']
            )
        await callback.answer(get_text(user_id, 'spec_updated'), show_alert=True)
    except Exception as e:
        logging.error(f"Spec update error: {e}")
        await callback.answer("Xatolik yuz berdi")

    await callback.message.delete()
    await show_profile(callback.message)

@dp.callback_query(F.data == "back_to_profile_view")
async def back_to_profile_view(callback: types.CallbackQuery):
    await callback.message.delete()
    await show_profile(callback.message)

@dp.callback_query(F.data == "edit_name")
async def edit_name_start(callback: types.CallbackQuery, state: FSMContext):
    user_id = callback.from_user.id
    
    await callback.answer()
    
    await state.set_state(ProfileEdit.waiting_for_new_name)
    await callback.message.edit_text(
        get_text(user_id, 'enter_new_name')
    )

@dp.message(ProfileEdit.waiting_for_new_name)
async def process_new_name(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    new_name = message.text.strip()
    
    if len(new_name) < 3:
        await message.answer("❌ Ism juda qisqa. Iltimos, qaytadan kiriting:")
        return
    
    user_names[user_id] = new_name
    
    try:
        user = await db.get_user(user_id)
        if user:
            await db.save_user(
                user_id=user_id,
                full_name=new_name,
                specialty=user['specialty'],
                language=user['language']
            )
    except Exception as e:
        logging.error(f"PostgreSQL da ism yangilashda xatolik: {e}")
    
    await state.clear()
    
    await message.answer(
        get_text(user_id, 'name_updated')
    )
    
    await show_profile(message)

@dp.callback_query(F.data == "back_to_main")
async def back_to_main_menu(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    
    await callback.answer()
    await callback.message.delete()
    
    keyboard = await main_keyboard(user_id)
    await callback.message.answer(
        "🏠 Asosiy menyu",
        reply_markup=keyboard
    )

@dp.message(F.text.in_({'📋 Dars jadvalim (PDF)', '📋 Мое расписание (PDF)', '📋 내 시간표 (PDF)'}))
async def view_my_schedule_pdf(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    if user_id not in user_schedules or not user_schedules[user_id]:
        await message.answer(get_text(user_id, 'no_schedules'))
        return
    
    try:
        pdf_buffer = await create_schedule_pdf(user_id)
        
        clean_name = clean_pdf_text(user_names.get(user_id, 'user'))
        await message.answer_document(
            types.BufferedInputFile(pdf_buffer.getvalue(), 
                                    filename=f"Dars_Jadvali_{clean_name}.pdf"),
            caption=get_text(user_id, 'my_schedule')
        )
    except Exception as e:
        logging.error(f"view_my_schedule_pdf error: {e}")
        await message.answer("❌ PDF yaratishda xatolik yuz berdi")

@dp.message(F.text.in_({'\U0001F4CA Mening statistikam', '\U0001F4CA Моя статистика', '\U0001F4CA 내 통계'}))
async def my_stats(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    now_uzb = datetime.now(UZB_TZ)
    current_month = now_uzb.strftime("%Y-%m")
    
    user_attendances = defaultdict(list)
    
    for (uid, branch, date, time) in daily_attendance_log:
        if uid == user_id:
            user_attendances[branch].append((date, time))
    
    if not user_attendances:
        await message.answer(get_text(user_id, 'no_stats'))
        return
    
    month_names_uz = {
        "01": "Yanvar", "02": "Fevral", "03": "Mart", "04": "Aprel",
        "05": "May", "06": "Iyun", "07": "Iyul", "08": "Avgust",
        "09": "Sentabr", "10": "Oktabr", "11": "Noyabr", "12": "Dekabr"
    }
    
    month_names_ru = {
        "01": "Январь", "02": "Февраль", "03": "Март", "04": "Апрель",
        "05": "Май", "06": "Июнь", "07": "Июль", "08": "Август",
        "09": "Сентябрь", "10": "Октябрь", "11": "Ноябрь", "12": "Декабрь"
    }
    
    month_names_kr = {
        "01": "1월", "02": "2월", "03": "3월", "04": "4월",
        "05": "5월", "06": "6월", "07": "7월", "08": "8월",
        "09": "9월", "10": "10월", "11": "11월", "12": "12월"
    }
    
    lang = user_languages.get(user_id, 'uz')
    if lang == 'uz':
        month_names = month_names_uz
        weekdays =["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]
        current_month_text = "(joriy oy)"
    elif lang == 'ru':
        month_names = month_names_ru
        weekdays =["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
        current_month_text = "(текущий месяц)"
    else:
        month_names = month_names_kr
        weekdays =["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]
        current_month_text = "(이번 달)"
    
    text = get_text(user_id, 'stats') + "\n\n"
    
    for branch, date_time_list in user_attendances.items():
        text += f"🏢 {branch}\n"
        
        dates_by_month = defaultdict(list)
        for date_str, time_str in date_time_list:
            year_month = date_str[:7]
            dates_by_month[year_month].append((date_str, time_str))
        
        for year_month, month_data in sorted(dates_by_month.items(), reverse=True):
            year, month = year_month.split('-')
            month_name = month_names.get(month, month)
            
            month_display = f"{month_name} {year}"
            if year_month == current_month:
                month_display += f" {current_month_text}"
            
            text += f"   📅 {month_display}\n"
            
            for date_str, time_str in sorted(month_data, reverse=True):
                date_parts = date_str.split('-')
                year, month, day = date_parts
                
                date_obj = datetime(int(year), int(month), int(day), tzinfo=UZB_TZ)
                weekday = date_obj.weekday()
                weekday_name = weekdays[weekday]
                
                if lang == 'kr':
                    formatted_date = f"{year}년 {int(month):02d}월 {int(day):02d}일"
                else:
                    formatted_date = f"{int(day):02d}.{int(month):02d}.{year}"
                
                text += f"      • {formatted_date} ({weekday_name}) - ⏰ {time_str}\n"
            
            text += "\n"
        
        text += "\n"
    
    await message.answer(text)

@dp.message(F.text.in_({'\U0001F3E2 Filiallar', '\U0001F3E2 Филиалы', '\U0001F3E2 지점'}))
async def show_branches(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    builder = InlineKeyboardBuilder()
    
    for loc in LOCATIONS:
        maps_link = get_yandex_maps_link(loc['lat'], loc['lon'])
        builder.row(
            InlineKeyboardButton(text=f"📍 {loc['name']}", url=maps_link)
        )
    
    await message.answer(
        "🏢 Mavjud filiallar (lokatsiya uchun bosing):",
        reply_markup=builder.as_markup()
    )

@dp.message(F.text.in_({'\u2753 Yordam', '\u2753 Помощь', '\u2753 도움말'}))
async def help_command(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    await message.answer(
        get_text(user_id, 'help')
    )

@dp.message(F.text.in_({'\U0001F3C6 Hafta topi', '\U0001F3C6 Топ недели', '\U0001F3C6 주간 TOP'}))
async def weekly_top(message: types.Message):
    user_id = message.from_user.id
    
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return
    
    now_uzb = datetime.now(UZB_TZ)
    week_ago = now_uzb - timedelta(days=7)
    week_ago_str = week_ago.strftime("%Y-%m-%d")
    
    weekly_stats = defaultdict(int)
    
    for (uid, branch, date, time) in daily_attendance_log:
        if date >= week_ago_str:
            weekly_stats[uid] += 1
    
    if not weekly_stats:
        lang = user_languages.get(user_id, 'uz')
        if lang == 'uz':
            no_data_msg = "💭 Bu hafta hali davomat yo'q"
        elif lang == 'ru':
            no_data_msg = "💭 На этой неделе еще нет отметок"
        else:
            no_data_msg = "💭 이번 주에는 아직 출석 기록이 없습니다"
        
        await message.answer(no_data_msg)
        return
    
    top_users = sorted(weekly_stats.items(), key=lambda x: x[1], reverse=True)[:10]
    
    top_list = ""
    for i, (uid, count) in enumerate(top_users, 1):
        try:
            name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            specialty_display = f" [{specialty}]" if specialty else ""
        except:
            name = f"Foydalanuvchi {uid}"
            specialty_display = ""
        
        medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
        top_list += f"{medal} {name}{specialty_display}: **{count}** marta\n"
    
    await message.answer(
        get_text(user_id, 'weekly_top', top_list=top_list)
    )

# --- BARQAROR LOKATSIYA HANDLERI (FAQAT OCHIQ FORWARD TEKSHIRILADI) ---
@dp.message(F.location)
async def handle_location(message: types.Message, state: FSMContext):
    user_id = message.from_user.id
    user_ids.add(user_id)
    
    # 1. BLOKLANGAN FOYDALANUVCHINI TEKSHIRISH
    if user_status.get(user_id) == 'blocked':
        await message.answer(get_text(user_id, 'blocked_user'))
        return

    # 2. FAQAT OCHIQ FORWARDNI TEKSHIRAMIZ
    if message.forward_origin is not None:
        user_warning = (
            "⚠️ **DIQQAT: SOXTA DAVOMATGA URINISH!**\n\n"
            "Siz boshqa foydalanuvchidan uzatilgan (forward) lokatsiyani yuborish orqali "
            "**yolg'on davomat** qilishga urundingiz.\n\n"
            "🚫 Ushbu harakatingiz soxtakorlik sifatida qayd etildi va adminlarga yuborildi!"
        )
        await message.answer(user_warning, parse_mode="Markdown")

        t_name = user_names.get(user_id, message.from_user.full_name)
        t_spec = user_specialty.get(user_id, 'Noma\'lum')
        admin_alert = (
            f"🚨 **SOXTA DAVOMATGA URINISH!**\n\n"
            f"👤 Xodim: {t_name}\n"
            f"📚 Soha: {t_spec}\n"
            f"🆔 ID: `{user_id}`\n"
            f"📍 Holat: Forward qilingan lokatsiya yubordi.\n"
            f"🕒 Vaqt: {datetime.now(UZB_TZ).strftime('%H:%M:%S')}"
        )
        await bot.send_message(ADMIN_GROUP_ID, admin_alert, parse_mode="Markdown")
        return

    # 3. MASOFANI O'LCHASH
    now_uzb = datetime.now(UZB_TZ)
    today_date = now_uzb.strftime("%Y-%m-%d")
    current_month = now_uzb.strftime("%Y-%m")
    now_time = now_uzb.strftime("%H:%M:%S")

    user_coords = (message.location.latitude, message.location.longitude)
    found_branch = None
    min_distance = float('inf')
    
    for branch in LOCATIONS:
        dist = geodesic((branch["lat"], branch["lon"]), user_coords).meters
        if dist <= ALLOWED_DISTANCE:
            if dist < min_distance:
                min_distance = dist
                found_branch = branch["name"]
    
    # 4. NATIJAGA QARAB JAVOB BERISH
    if found_branch:
        # Avval davomat qilganmi?
        already_attended = any(k[0] == user_id and k[1] == found_branch and k[2] == today_date for k in daily_attendance_log)
        if already_attended:
            await message.answer(get_text(user_id, 'already_attended', branch=found_branch), parse_mode="Markdown")
            return

        # PostgreSQL va RAM ga saqlash
        await db.save_attendance(user_id, found_branch, today_date, now_time)
        daily_attendance_log.add((user_id, found_branch, today_date, now_time))
        
        counter_key = (user_id, found_branch, current_month)
        attendance_counter[counter_key] = attendance_counter.get(counter_key, 0) + 1
        visit_number = attendance_counter[counter_key]
        
        # Adminga hisobot
        full_name = user_names.get(user_id, message.from_user.full_name)
        specialty = user_specialty.get(user_id, '')
        specialty_display = f" [{specialty}]" if specialty else ""
        
        report = (
            f"✅ **Yangi Davomat**\n\n"
            f"👤 O'qituvchi: {full_name}{specialty_display}\n"
            f"📍 Manzil: {found_branch}\n"
            f"📅 Sana: {today_date}\n"
            f"⏰ Vaqt: {now_time}\n"
            f"📊 Shu oydagi tashrif: {visit_number}-marta\n"
            f"📍 Masofa: {min_distance:.1f} metr"
        )
        await bot.send_message(chat_id=ADMIN_GROUP_ID, text=report, parse_mode="Markdown")

        # Foydalanuvchiga muvaffaqiyat xabari
        success_text = get_text(
            user_id, 
            'attendance_success', 
            branch=found_branch, 
            date=today_date, 
            time=now_time, 
            count=visit_number, 
            distance=min_distance
        )
        
        # Ob-havo ma'lumotini qo'shamiz
        weather_data = await get_weather_by_coords(user_coords[0], user_coords[1])
        weather_message = format_weather_message(weather_data, user_languages.get(user_id, 'uz'))
        
        await message.answer(f"{success_text}\n\n{weather_message}", parse_mode="Markdown")

        # --- YANGI: O'QUVCHILAR DAVOMATI TIZIMI ---
        # Shu o'qituvchining shu filialdagi guruhini qidiramiz
        current_grp = None
        for g_id, g_data in groups.items():
            if g_data['teacher_id'] == user_id and g_data['branch'] == found_branch:
                # Agar bugun dars kuni bo'lsa
                if WEEKDAYS_UZ[now_uzb.weekday()] in g_data['days']:
                    current_grp = g_id
                    break
        
        if current_grp:
            await state.update_data(current_group_id=current_grp, selected_stds=[])
            keyboard = await get_student_attendance_kb(current_grp, [])
            await message.answer(
                f"🧑‍🎓 **{groups[current_grp]['group_name']}** guruhi o'quvchilari.\n"
                f"Darsda qatnashayotganlarni belgilang:",
                reply_markup=keyboard, parse_mode="Markdown"
            )
            await state.set_state(StudentAttendance.selecting_students)
    else:
        # FAQAT HUDUDDAN TASHQARIDA BO'LSA
        await message.answer(get_text(user_id, 'not_in_area'), parse_mode="Markdown")

# --- O'QUVCHILAR DAVOMATI HANDLERLARI ---
@dp.callback_query(StudentAttendance.selecting_students, F.data.startswith("std_check_"))
async def std_check_callback(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.replace("std_check_", ""))
    data = await state.get_data()
    selected = data.get('selected_stds', [])
    
    if idx in selected:
        selected.remove(idx)
    else:
        selected.append(idx)
    
    await state.update_data(selected_stds=selected)
    await callback.message.edit_reply_markup(
        reply_markup=await get_student_attendance_kb(data['current_group_id'], selected)
    )
    await callback.answer()

@dp.callback_query(StudentAttendance.selecting_students, F.data == "std_submit")
async def std_submit_callback(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    g_id = data['current_group_id']
    selected = data.get('selected_stds', [])
    students = group_students.get(g_id, [])
    user_id = callback.from_user.id
    
    # Excel yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Guruh Davomati"
    
    # Sarlavhalar
    ws.append(['№', 'O\'quvchi', 'Telefon', 'Holat'])
    
    # O'quvchilar ro'yxati
    for i, s in enumerate(students):
        status = "Kelgan" if i in selected else "Kelmagan"
        ws.append([i+1, s['name'], s['phone'], status])
    
    # Ustun kengliklarini sozlash
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    
    # Formatlash
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2, max_row=len(students)+1):
        for cell in row:
            cell.border = border
    
    # Sarlavha formatlash
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    # Adminga yuborish
    group_name = groups[g_id]['group_name']
    teacher_name = user_names.get(user_id, 'Noma\'lum')
    current_date = datetime.now(UZB_TZ).strftime('%Y-%m-%d')
    
    caption = (
        f"🧑‍🎓 **Guruh Davomati**\n"
        f"📦 Guruh: {group_name}\n"
        f"👤 O'qituvchi: {teacher_name}\n"
        f"📅 Sana: {current_date}\n"
        f"👥 Kelganlar: {len(selected)}/{len(students)}"
    )
    
    filename = f"Guruh_Davomati_{group_name}_{current_date}.xlsx"
    
    await bot.send_document(
        ADMIN_GROUP_ID, 
        types.BufferedInputFile(buf.read(), filename=filename),
        caption=caption,
        parse_mode="Markdown"
    )
    
    # O'qituvchiga javob
    await callback.message.edit_text(
        f"✅ O'quvchilar davomati yuborildi!\n\n"
        f"Kelganlar: {len(selected)}/{len(students)}"
    )
    await state.clear()

async def get_weather_by_coords(lat: float, lon: float):
    params = {
        "lat": lat,
        "lon": lon,
        "appid": WEATHER_API_KEY,
        "units": "metric",
        "lang": "uz"
    }
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(WEATHER_API_URL, params=params) as response:
                if response.status == 200:
                    data = await response.json()
                    return data
                else:
                    logging.error(f"Weather API error: {response.status}")
                    return None
    except Exception as e:
        logging.error(f"Error fetching weather: {e}")
        return None

def get_weather_emoji(weather_condition: str) -> str:
    emoji_map = {
        "Clear": "☀️",
        "Clouds": "☁️",
        "Rain": "🌧️",
        "Drizzle": "🌦️",
        "Thunderstorm": "⛈️",
        "Snow": "❄️",
        "Mist": "🌫️",
        "Fog": "🌫️",
        "Haze": "🌫️"
    }
    return emoji_map.get(weather_condition, "🌡️")

def format_weather_message(weather_data: dict, lang: str = 'uz') -> str:
    if not weather_data:
        return ""
    
    city = weather_data.get('name', 'Noma\'lum')
    if city == "" or city is None:
        city = "Toshkent"
        
    main = weather_data.get('main', {})
    weather = weather_data.get('weather', [{}])[0]
    wind = weather_data.get('wind', {})
    
    temp = main.get('temp', 0)
    feels_like = main.get('feels_like', 0)
    humidity = main.get('humidity', 0)
    pressure = main.get('pressure', 0)
    condition = weather.get('main', 'Unknown')
    description = weather.get('description', '')
    wind_speed = wind.get('speed', 0)
    
    emoji = get_weather_emoji(condition)
    
    recommendation = WEATHER_RECOMMENDATIONS.get(condition, {}).get(lang, "")
    if not recommendation:
        recommendation = WEATHER_RECOMMENDATIONS.get('Clear', {}).get(lang, "")
    
    pressure_mmhg = pressure * 0.750062
    
    temp_text = "Harorat" if lang == 'uz' else "Температура" if lang == 'ru' else "기온"
    feels_text = "his qilinadi" if lang == 'uz' else "ощущается" if lang == 'ru' else "체감"
    humidity_text = "Namlik" if lang == 'uz' else "Влажность" if lang == 'ru' else "습도"
    wind_text = "Shamol" if lang == 'uz' else "Ветер" if lang == 'ru' else "바람"
    pressure_text = "Bosim" if lang == 'uz' else "Давление" if lang == 'ru' else "기압"
    recommendation_title = "Tavsiya" if lang == 'uz' else "Рекомендация" if lang == 'ru' else "추천"
    time_text = "Vaqt" if lang == 'uz' else "Время" if lang == 'ru' else "시간"
    
    message = f"""
{emoji} Ob-havo ma'lumoti

📍 Joy: {city}
🌡️ {temp_text}: {temp:.1f}°C ({feels_text}: {feels_like:.1f}°C)
💧 {humidity_text}: {humidity}%
💨 {wind_text}: {wind_speed:.1f} m/s
📊 {pressure_text}: {pressure_mmhg:.1f} mmHg

💡 {recommendation_title}:
{recommendation}

⏰ {time_text}: {datetime.now(UZB_TZ).strftime('%H:%M')}
"""
    return message

async def get_calendar_keyboard(year: int, month: int, lang: str):
    builder = InlineKeyboardBuilder()
    
    month_names = {
        'uz': ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun", "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"],
        'ru': ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"],
        'kr': ["1월", "2월", "3월", "4월", "5월", "6월", "7월", "8월", "9월", "10월", "11월", "12월"]
    }
    
    wd_names = {
        'uz': ["Du", "Se", "Ch", "Pa", "Ju", "Sha", "Ya"],
        'ru': ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"],
        'kr': ["월", "화", "수", "목", "금", "토", "일"]
    }

    m_name = month_names[lang][month-1]
    header_text = f"{m_name} {year}" if lang != 'kr' else f"{year}년 {m_name}"
    
    builder.row(
        InlineKeyboardButton(text="⬅️", callback_data=f"cal_nav_prev_{year}_{month}"),
        InlineKeyboardButton(text=header_text, callback_data="ignore"),
        InlineKeyboardButton(text="➡️", callback_data=f"cal_nav_next_{year}_{month}")
    )

    header_days = [InlineKeyboardButton(text=day, callback_data="ignore") for day in wd_names[lang]]
    builder.row(*header_days)

    month_calendar = calendar.monthcalendar(year, month)
    for week in month_calendar:
        row_btns = []
        for day in week:
            if day == 0:
                row_btns.append(InlineKeyboardButton(text=" ", callback_data="ignore"))
            else:
                row_btns.append(InlineKeyboardButton(
                    text=str(day), 
                    callback_data=f"cal_set_{year}-{month:02d}-{day:02d}")
                )
        builder.row(*row_btns)

    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    return builder.as_markup()

@dp.callback_query(F.data == "admin_pdf_report")
async def admin_pdf_report_start(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    now = datetime.now(UZB_TZ)
    lang = user_languages.get(callback.from_user.id, 'uz')
    
    keyboard = await get_calendar_keyboard(now.year, now.month, lang)
    await callback.message.edit_text(
        "📅 Hisobot sanasini kalendardan tanlang:",
        reply_markup=keyboard
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("cal_nav_"))
async def process_calendar_navigation(callback: types.CallbackQuery):
    parts = callback.data.split("_")
    action = parts[2]
    year = int(parts[3])
    month = int(parts[4])
    
    if action == "prev":
        month -= 1
        if month == 0:
            month = 12
            year -= 1
    else:
        month += 1
        if month == 13:
            month = 1
            year += 1
            
    lang = user_languages.get(callback.from_user.id, 'uz')
    keyboard = await get_calendar_keyboard(year, month, lang)
    
    await callback.message.edit_reply_markup(reply_markup=keyboard)
    await callback.answer()

@dp.callback_query(F.data.startswith("cal_set_"))
async def process_calendar_selection(callback: types.CallbackQuery):
    date_str = callback.data.replace("cal_set_", "")
    report_date = datetime.strptime(date_str, "%Y-%m-%d").date()
    
    await callback.message.answer(f"⏳ {date_str} davomat hisoboti tayyorlanmoqda...")
    
    try:
        pdf_buf = await get_combined_report_pdf(report_date)
        
        await callback.message.answer_document(
            types.BufferedInputFile(pdf_buf.read(), filename=f"Davomat_{date_str}.pdf"),
            caption=f"📊 {date_str} hisoboti (PDF formatda)."
        )
    except Exception as e:
        logging.error(f"Calendar PDF error: {e}")
        await callback.message.answer("❌ PDF yaratishda xatolik yuz berdi.")
        
    await callback.answer()

@dp.callback_query(F.data == "ignore")
async def process_ignore_callback(callback: types.CallbackQuery):
    await callback.answer()

@dp.message(Command("admin"))
async def admin_panel(message: types.Message):
    if not check_admin(message.chat.id):
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="💰 Oylik hisoblash", callback_data="admin_salary_calc")
        )
        builder.row(
            InlineKeyboardButton(text="🖼 Vizual Jadval (Haftalik)", callback_data="admin_visual_schedule")
        )
        builder.row(
            InlineKeyboardButton(text="➕ Guruh shakllantirish", callback_data="admin_create_group")
        )
        builder.row(
            InlineKeyboardButton(text="👥 Foydalanuvchilar", callback_data="admin_users_main"),
            InlineKeyboardButton(text="📢 Xabar yuborish", callback_data="admin_broadcast")
        )
        builder.row(
            InlineKeyboardButton(text="🏢 Filiallar", callback_data="admin_locations_main"),
            InlineKeyboardButton(text="📅 Dars jadvallari", callback_data="admin_schedules_main")
        )
        builder.row(
            InlineKeyboardButton(text="📊 Oylik hisobot (Excel)", callback_data="admin_excel_menu"),
            InlineKeyboardButton(text="📊 Kunlik PDF", callback_data="admin_pdf_report")
        )
        
        await message.answer(
            "👨‍💼 Admin Panel\n\nKerakli bo'limni tanlang:",
            reply_markup=builder.as_markup()
        )
    except Exception as e:
        logging.error(f"admin_panel error: {e}")
        await message.answer("❌ Admin panelni ochishda xatolik yuz berdi")

# --- GURUH YARATISH HANDLERLARI (YANGI) - TUZATILGAN VERSIYA---
@dp.callback_query(F.data == "admin_create_group")
async def start_group_creation(callback: types.CallbackQuery, state: FSMContext):
    builder = InlineKeyboardBuilder()
    for loc in LOCATIONS:
        builder.row(InlineKeyboardButton(text=loc['name'], callback_data=f"grp_br_{loc['name']}"))
    builder.row(InlineKeyboardButton(text="🔙 Bekor qilish", callback_data="admin_back"))
    await callback.message.edit_text("Qaysi filialga guruh qo'shmoqchisiz?", reply_markup=builder.as_markup())
    await state.set_state(CreateGroup.selecting_branch)

@dp.callback_query(CreateGroup.selecting_branch, F.data.startswith("grp_br_"))
async def grp_branch_selected(callback: types.CallbackQuery, state: FSMContext):
    await state.update_data(branch=callback.data.replace("grp_br_", ""))
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="💻 IT", callback_data="grp_type_IT"),
                InlineKeyboardButton(text="🇰🇷 Koreys tili", callback_data="grp_type_Koreys tili"))
    await callback.message.edit_text("Dars turini tanlang:", reply_markup=builder.as_markup())
    await state.set_state(CreateGroup.selecting_type)

@dp.callback_query(CreateGroup.selecting_type, F.data.startswith("grp_type_"))
async def grp_type_selected(callback: types.CallbackQuery, state: FSMContext):
    await state.update_data(type=callback.data.replace("grp_type_", ""))
    builder = InlineKeyboardBuilder()
    # Ofis xodimlari bundan mustasno (faqat o'qituvchilar)
    for uid, name in user_names.items():
        if user_specialty.get(uid) != 'Ofis xodimi' and user_specialty.get(uid) is not None:
            builder.row(InlineKeyboardButton(text=name, callback_data=f"grp_tchr_{uid}"))
    await callback.message.edit_text("O'qituvchini tanlang:", reply_markup=builder.as_markup())
    await state.set_state(CreateGroup.selecting_teacher)

@dp.callback_query(CreateGroup.selecting_teacher, F.data.startswith("grp_tchr_"))
async def grp_teacher_selected(callback: types.CallbackQuery, state: FSMContext):
    teacher_id = int(callback.data.replace("grp_tchr_", ""))
    await state.update_data(teacher_id=teacher_id, selected_days=[])
    await state.set_state(CreateGroup.selecting_days)
    await grp_show_days(callback.message, [])
    await callback.answer()

async def grp_show_days(message: types.Message, selected):
    builder = InlineKeyboardBuilder()
    for day in WEEKDAYS_UZ:
        text = f"✅ {day}" if day in selected else f"⬜ {day}"
        builder.row(InlineKeyboardButton(text=text, callback_data=f"grp_day_{day}"))
    builder.row(InlineKeyboardButton(text="➡️ Keyingisi", callback_data="grp_days_next"))
    await message.edit_text("Dars kunlarini tanlang (bir nechta tanlashingiz mumkin):", reply_markup=builder.as_markup())

@dp.callback_query(CreateGroup.selecting_days, F.data.startswith("grp_day_"))
async def grp_toggle_day(callback: types.CallbackQuery, state: FSMContext):
    day = callback.data.replace("grp_day_", "")
    data = await state.get_data()
    selected = data.get('selected_days', [])
    if day in selected:
        selected.remove(day)
    else:
        selected.append(day)
    await state.update_data(selected_days=selected)
    await grp_show_days(callback.message, selected)
    await callback.answer()

@dp.callback_query(CreateGroup.selecting_days, F.data == "grp_days_next")
async def grp_days_next(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    if not data.get('selected_days'):
        await callback.answer("Hech bo'lmaganda 1 kun tanlang!", show_alert=True)
        return
    await callback.message.edit_text("Dars soatini kiriting (masalan, 14:00):")
    await state.set_state(CreateGroup.entering_time)

@dp.message(CreateGroup.entering_time)
async def grp_time_entered(message: types.Message, state: FSMContext):
    # Vaqt formatini tekshirish
    time_pattern = re.compile(r'^([0-1]?[0-9]|2[0-3]):[0-5][0-9]$')
    if not time_pattern.match(message.text.strip()):
        await message.answer("❌ Noto'g'ri format! Iltimos, HH:MM formatida kiriting (masalan, 14:00):")
        return
    await state.update_data(time=message.text.strip())
    await message.answer("Guruh uchun nom kiriting (masalan, Koreys-1 yoki IT-A):")
    await state.set_state(CreateGroup.entering_group_name)

@dp.message(CreateGroup.entering_group_name)
async def grp_name_entered(message: types.Message, state: FSMContext):
    await state.update_data(group_name=message.text.strip(), students=[])
    await message.answer(
        "Endi o'quvchilarni qo'shishni boshlaymiz.\n\n"
        "1-o'quvchining ismi va familiyasini kiriting:"
    )
    await state.set_state(CreateGroup.adding_student_name)

@dp.message(CreateGroup.adding_student_name)
async def grp_student_name(message: types.Message, state: FSMContext):
    if len(message.text.strip()) < 2:
        await message.answer("❌ Ism juda qisqa. Qaytadan kiriting:")
        return
    await state.update_data(current_std_name=message.text.strip())
    await message.answer(f"'{message.text.strip()}' ning telefon raqamini kiriting:")
    await state.set_state(CreateGroup.adding_student_phone)

@dp.message(CreateGroup.adding_student_phone)
async def grp_student_phone(message: types.Message, state: FSMContext):
    phone = message.text.strip()
    data = await state.get_data()
    students = data.get('students', [])
    students.append({"name": data['current_std_name'], "phone": phone})
    await state.update_data(students=students)
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="➕ Yana o'quvchi qo'shish", callback_data="grp_add_more"),
        InlineKeyboardButton(text="✅ Guruhni yakunlash", callback_data="grp_finish")
    )
    await message.answer(
        f"✅ O'quvchi qo'shildi! Jami: {len(students)} ta.\n\n"
        f"Davom etamizmi?",
        reply_markup=builder.as_markup()
    )
    await state.set_state(CreateGroup.confirm_student)

@dp.callback_query(CreateGroup.confirm_student, F.data == "grp_add_more")
async def grp_more_students(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    await callback.message.edit_text(
        f"{len(data['students'])+1}-o'quvchi ismini kiriting:"
    )
    await state.set_state(CreateGroup.adding_student_name)

# --- TUZATILGAN GURUH YARATISH FUNKSIYASI (f-string xatosi tuzatildi) ---
@dp.callback_query(CreateGroup.confirm_student, F.data == "grp_finish")
async def grp_final_save(callback: types.CallbackQuery, state: FSMContext):
    data = await state.get_data()
    
    # 1. Bazaga saqlash
    try:
        async with db.pool.acquire() as conn:
            group_id = await conn.fetchval("""
                INSERT INTO groups (group_name, branch, lesson_type, teacher_id, days_data, time_text)
                VALUES ($1, $2, $3, $4, $5::jsonb, $6) RETURNING id
            """, data['group_name'], data['branch'], data['type'], data['teacher_id'], 
               json.dumps(data['selected_days']), data['time'])
            
            for std in data['students']:
                await conn.execute("""
                    INSERT INTO group_students (group_id, student_name, student_phone) 
                    VALUES ($1, $2, $3)
                """, group_id, std['name'], std['phone'])
        
        # RAM ga qo'shish
        groups[group_id] = {
            'group_name': data['group_name'],
            'branch': data['branch'],
            'lesson_type': data['type'],
            'teacher_id': data['teacher_id'],
            'days': data['selected_days'],
            'time': data['time']
        }
        group_students[group_id] = data['students']

        # 2. O'qituvchiga xabar - f-string ichida murakkab amal yo'q
        days_str = ", ".join(data['selected_days'])  # JOIN AMALI ALOHIDA BAJARILDI
        teacher_msg = (
            f"🆕 **Yangi guruh biriktirildi!**\n\n"
            f"📦 Guruh: {data['group_name']}\n"
            f"🏢 Filial: {data['branch']}\n"
            f"📚 Fan: {data['type']}\n"
            f"⏰ Vaqt: {data['time']}\n"
            f"📅 Kunlar: {days_str}\n"
            f"👥 O'quvchilar soni: {len(data['students'])} ta\n\n"
            f"Botda davomat qilganingizda ushbu o'quvchilar ro'yxati chiqadi."
        )
        try:
            await bot.send_message(data['teacher_id'], teacher_msg, parse_mode="Markdown")
        except Exception as e:
            logging.error(f"Failed to notify teacher: {e}")

        await callback.message.edit_text(
            f"✅ Guruh '{data['group_name']}' muvaffaqiyatli yaratildi!\n\n"
            f"🏢 Filial: {data['branch']}\n"
            f"📚 Fan: {data['type']}\n"
            f"👤 O'qituvchi: {user_names.get(data['teacher_id'])}\n"
            f"👥 O'quvchilar: {len(data['students'])} ta"
        )
    except Exception as e:
        logging.error(f"Error saving group: {e}")
        await callback.message.edit_text(f"❌ Xatolik yuz berdi: {e}")
    
    await state.clear()

async def create_visual_timetable_img(branch_name: str):
    days = ['Dushanba', 'Seshanba', 'Chorshanba', 'Payshanba', 'Juma', 'Shanba', 'Yakshanba']
    time_slots = ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00', '18:00', '19:00', '20:00']
    
    plt.figure(figsize=(14, 8))
    ax = plt.gca()
    
    colors_map = {
        'IT': '#E3F2FD',
        'Koreys tili': '#E8F5E9',
        'Ofis xodimi': '#FFF3E0'
    }
    border_map = {
        'IT': '#1565C0', 
        'Koreys tili': '#2E7D32', 
        'Ofis xodimi': '#EF6C00'
    }

    for i in range(len(days) + 1):
        plt.axvline(i, color='gray', linestyle='--', alpha=0.3)
    for i in range(len(time_slots) + 1):
        plt.axhline(i, color='gray', linestyle='--', alpha=0.3)

    found_any = False
    for sid, data in schedules.items():
        if data['branch'] == branch_name:
            found_any = True
            uid = data['user_id']
            spec = user_specialty.get(uid, 'IT')
            t_name = user_names.get(uid, "Noma'lum")
            
            for day, t_val in data['days'].items():
                if day in days:
                    day_idx = days.index(day)
                    try:
                        h, m = map(int, t_val.split(':'))
                        start_y = h + (m/60)
                        y_pos = len(time_slots) - (start_y - 8)
                        
                        rect = plt.Rectangle((day_idx + 0.05, y_pos - 0.9), 0.9, 0.8, 
                                            facecolor=colors_map.get(spec, '#F5F5F5'),
                                            edgecolor=border_map.get(spec, 'gray'),
                                            linewidth=1.5, alpha=0.9, zorder=3)
                        ax.add_patch(rect)
                        
                        plt.text(day_idx + 0.5, y_pos - 0.5, f"{t_name}\n({t_val})\n{spec}", 
                                 ha='center', va='center', fontsize=8, fontweight='bold', zorder=4)
                    except: continue

    plt.xticks(np.arange(0.5, len(days), 1), days, fontweight='bold')
    plt.yticks(np.arange(0.5, len(time_slots), 1), time_slots[::-1], fontweight='bold')
    
    plt.title(f"🏢 {branch_name} - Haftalik Bandlik Jadvali", fontsize=16, pad=20, fontweight='bold', color='#1A237E')
    plt.xlim(0, len(days))
    plt.ylim(0, len(time_slots))
    
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0], marker='s', color='w', label='IT Bo\'limi', markerfacecolor='#E3F2FD', markersize=15, markeredgecolor='#1565C0'),
        Line2D([0], [0], marker='s', color='w', label='Koreys tili', markerfacecolor='#E8F5E9', markersize=15, markeredgecolor='#2E7D32'),
        Line2D([0], [0], marker='s', color='w', label='Ofis xodimi', markerfacecolor='#FFF3E0', markersize=15, markeredgecolor='#EF6C00')
    ]
    ax.legend(handles=legend_elements, loc='upper center', bbox_to_anchor=(0.5, -0.05), ncol=3)

    img_buf = io.BytesIO()
    plt.savefig(img_buf, format='png', bbox_inches='tight', dpi=150)
    img_buf.seek(0)
    plt.close()
    return img_buf, found_any

@dp.callback_query(F.data == "admin_visual_schedule")
async def visual_schedule_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    builder = InlineKeyboardBuilder()
    for loc in LOCATIONS:
        builder.row(InlineKeyboardButton(text=loc['name'], callback_data=f"v_br_{loc['name']}"))
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    
    await callback.message.edit_text(
        "Qaysi filialning vizual bandlik jadvalini ko'rmoqchisiz?",
        reply_markup=builder.as_markup()
    )
    await state.set_state(VisualSchedule.selecting_branch)
    await callback.answer()

@dp.callback_query(VisualSchedule.selecting_branch, F.data.startswith("v_br_"))
async def visual_schedule_process(callback: types.CallbackQuery, state: FSMContext):
    branch_name = callback.data.replace("v_br_", "")
    await callback.message.answer(f"⏳ {branch_name} uchun vizual jadval tayyorlanmoqda...")
    
    try:
        img_buf, found = await create_visual_timetable_img(branch_name)
        if not found:
            await callback.message.answer(f"📭 {branch_name} filialida hali darslar belgilanmagan.")
        else:
            await callback.message.answer_photo(
                types.BufferedInputFile(img_buf.read(), filename="timetable.png"),
                caption=f"🖼 {branch_name} filialining haftalik bandlik xaritasi."
            )
    except Exception as e:
        logging.error(f"Visual schedule error: {e}")
        traceback.print_exc()
        await callback.message.answer("❌ Jadvalni chizishda xatolik yuz berdi.")
    
    await state.clear()
    await callback.answer()

# --- OYLIK KALKULYATOR HANDLERS (TUZATILGAN VERSIYA)---
@dp.callback_query(F.data == "admin_salary_calc")
async def salary_calc_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="💻 IT", callback_data="sal_spec_IT"),
        InlineKeyboardButton(text="🇰🇷 Koreys tili", callback_data="sal_spec_Koreys tili")
    )
    builder.row(
        InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
    )
    await callback.message.edit_text(
        "Qaysi fan o'qituvchisiga oylik hisoblamoqchisiz?",
        reply_markup=builder.as_markup()
    )
    await state.set_state(SalaryCalc.selecting_specialty)
    await callback.answer()

@dp.callback_query(SalaryCalc.selecting_specialty, F.data.startswith("sal_spec_"))
async def salary_calc_spec(callback: types.CallbackQuery, state: FSMContext):
    spec = callback.data.replace("sal_spec_", "")
    await state.update_data(specialty=spec)
    
    builder = InlineKeyboardBuilder()
    for uid, name in user_names.items():
        if user_specialty.get(uid) == spec:
            builder.row(InlineKeyboardButton(text=str(name), callback_data=f"sal_teacher_{uid}"))
    
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_salary_calc"))
    await callback.message.edit_text(
        f"{spec} o'qituvchisini tanlang:",
        reply_markup=builder.as_markup()
    )
    await state.set_state(SalaryCalc.selecting_teacher)
    await callback.answer()

@dp.callback_query(SalaryCalc.selecting_teacher, F.data.startswith("sal_teacher_"))
async def salary_calc_teacher_selected(callback: types.CallbackQuery, state: FSMContext):
    uid = int(callback.data.replace("sal_teacher_", ""))
    
    # O'qituvchining barcha filiallarini topamiz
    branches = []
    for sid, sdata in schedules.items():
        if sdata['user_id'] == uid:
            if sdata['branch'] not in branches:
                branches.append(sdata['branch'])
    
    if not branches:
        await callback.message.answer("❌ Bu o'qituvchiga dars jadvali biriktirilmagan!")
        await state.clear()
        return

    await state.update_data(
        teacher_id=uid,
        teacher_name=user_names.get(uid, f"ID: {uid}"),
        specialty=user_specialty.get(uid, ''),
        all_branches=branches,
        current_branch_idx=0,
        calculated_results=[]
    )
    
    await salary_ask_next_branch(callback.message, state)
    await callback.answer()

async def salary_ask_next_branch(message: types.Message, state: FSMContext):
    data = await state.get_data()
    idx = data['current_branch_idx']
    branches = data['all_branches']
    current_branch = branches[idx]
    
    await message.answer(f"🏢 Filial: {current_branch}\n\nUshbu filialdagi o'quvchilar sonini kiriting:")
    await state.set_state(SalaryCalc.entering_students)

@dp.message(SalaryCalc.entering_students)
async def salary_students_step(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Raqam kiriting!")
        return
    await state.update_data(temp_students=int(message.text))
    await message.answer("Ushbu filialda bu oy necha marta dars o'tdi?")
    await state.set_state(SalaryCalc.entering_lessons)

@dp.message(SalaryCalc.entering_lessons)
async def salary_lessons_step(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Raqam kiriting!")
        return
    await state.update_data(temp_lessons=int(message.text))
    
    data = await state.get_data()
    builder = InlineKeyboardBuilder()
    if data['specialty'] == "IT":
        builder.row(InlineKeyboardButton(text="35%", callback_data="p_it_35"),
                    InlineKeyboardButton(text="45%", callback_data="p_it_45"))
        await message.answer("Imtixon natijasini tanlang:", reply_markup=builder.as_markup())
    else:
        for p in range(10, 101, 10):
            builder.add(InlineKeyboardButton(text=f"{p}%", callback_data=f"p_kr_{p}"))
        builder.adjust(3)
        await message.answer("Imtixon natijasini tanlang (%):", reply_markup=builder.as_markup())
    await state.set_state(SalaryCalc.selecting_percentage)

@dp.callback_query(SalaryCalc.selecting_percentage)
async def salary_perc_step(callback: types.CallbackQuery, state: FSMContext):
    perc = int(callback.data.split('_')[-1])
    await state.update_data(temp_perc=perc)
    
    data = await state.get_data()
    if data['specialty'] == "IT":
        await callback.message.edit_text("Ushbu filial uchun jarima FOIZINI kiriting (masalan: 10. Jarima bo'lmasa 0):")
        await state.set_state(SalaryCalc.entering_penalty_it_percent)
    else:
        await callback.message.edit_text("Ushbu filial uchun jarima SUMMASINI kiriting (so'mda, masalan: 50000. Jarima bo'lmasa 0):")
        await state.set_state(SalaryCalc.entering_penalty_kr_sum)
    await callback.answer()

@dp.message(SalaryCalc.entering_penalty_it_percent)
async def salary_it_penalty_percent(message: types.Message, state: FSMContext):
    if not message.text.isdigit():
        await message.answer("Raqamda kiriting!")
        return
    await state.update_data(temp_penalty_val=int(message.text))
    await message.answer("Ushbu filialdan jami o'quvchilar to'lovini kiriting:")
    await state.set_state(SalaryCalc.entering_payment)

@dp.message(SalaryCalc.entering_penalty_kr_sum)
async def salary_kr_penalty_sum(message: types.Message, state: FSMContext):
    val = message.text.replace(' ', '').replace(',', '')
    if not val.isdigit():
        await message.answer("Raqamda kiriting!")
        return
    await state.update_data(temp_penalty_val=int(val))
    await process_branch_calculation(message, state)

@dp.message(SalaryCalc.entering_payment)
async def salary_payment_it_step(message: types.Message, state: FSMContext):
    val = message.text.replace(' ', '').replace(',', '')
    if not val.isdigit():
        await message.answer("Raqam kiriting!")
        return
    await state.update_data(temp_payment=int(val))
    await process_branch_calculation(message, state)

async def process_branch_calculation(message: types.Message, state: FSMContext):
    data = await state.get_data()
    spec = data['specialty']
    idx = data['current_branch_idx']
    branch_name = data['all_branches'][idx]
    
    gross = 0
    exam_pen = 0
    penalty_disp = ""

    if spec == "IT":
        share_amount = (data['temp_payment'] * data['temp_perc'] / 100)
        penalty_amount = (share_amount * data['temp_penalty_val'] / 100)
        gross = share_amount - penalty_amount
        penalty_disp = f"{data['temp_penalty_val']}%"
    else:
        students = data['temp_students']
        lessons = data['temp_lessons']
        perc = data['temp_perc']
        base = 1800000 + (students * 100000 if students > 10 else 0)
        
        exam_pen = get_kr_exam_penalty(perc)
        mid_total = base - exam_pen
        
        if lessons < 12:
            gross = (mid_total / 12) * lessons
        else:
            gross = mid_total
        
        gross -= data['temp_penalty_val']
        # Formatlashni alohida bajaramiz
        penalty_val = data['temp_penalty_val']
        penalty_disp = f"{penalty_val:,} so'm".replace(',', ' ')

    res = {
        'branch': branch_name,
        'students': data['temp_students'],
        'lessons': data['temp_lessons'],
        'perc': data['temp_perc'],
        'penalty_display': penalty_disp,
        'exam_penalty': exam_pen,
        'payment': data.get('temp_payment', 0),
        'gross': gross
    }
    
    results_list = data['calculated_results']
    results_list.append(res)
    
    new_idx = idx + 1
    if new_idx < len(data['all_branches']):
        await state.update_data(current_branch_idx=new_idx, calculated_results=results_list)
        await salary_ask_next_branch(message, state)
    else:
        await finalize_multi_branch_salary(message, state, results_list)

async def finalize_multi_branch_salary(message: types.Message, state: FSMContext, results):
    data = await state.get_data()
    total_gross = sum(r['gross'] for r in results)
    tax = total_gross * 0.075
    net = total_gross - tax
    
    excel_file = await create_multi_branch_excel(data['teacher_name'], data['specialty'], results, total_gross, tax, net)
    
    # Formatlashni alohida bajaramiz - f-string ichida replace ishlatilmadi
    s_net = "{:,.0f}".format(net).replace(',', ' ')
    s_tax = "{:,.0f}".format(tax).replace(',', ' ')
    s_gross = "{:,.0f}".format(total_gross).replace(',', ' ')
    
    caption = (f"💰 Hisob-kitob yakunlandi\n\n"
               f"👤 Xodim: {data['teacher_name']}\n"
               f"📚 Mutaxassislik: {data['specialty']}\n"
               f"🏢 Filiallar: {len(results)} ta\n"
               f"──────────────────\n"
               f"Jami (soliqsiz): {s_gross} so'm\n"
               f"Soliq (7.5%): {s_tax} so'm\n"
               f"Qo'lga tegadi: {s_net} so'm")
    
    await message.answer_document(
        types.BufferedInputFile(excel_file.read(), filename=f"Oylik_{data['teacher_name']}_{datetime.now(UZB_TZ).strftime('%Y%m')}.xlsx"),
        caption=caption
    )
    await state.clear()
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🔄 Yangi hisoblash", callback_data="admin_salary_calc"))
    builder.row(InlineKeyboardButton(text="🔙 Admin Panel", callback_data="admin_back"))
    await message.answer(
        "Boshqa amalni tanlang:",
        reply_markup=builder.as_markup()
    )

async def create_multi_branch_excel(teacher_name, specialty, results, total_gross, tax, net):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Oylik Hisoboti"
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    headers = ['Filial', 'O\'quvchilar', 'Darslar', 'Imtixon %', 'Jarima', 'Tushum', 'Hisoblangan oylik']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for r in results:
        row = [
            r['branch'],
            r['students'],
            r['lessons'],
            f"{r['perc']}%",
            r['penalty_display'],
            f"{r['payment']:,.0f}" if r['payment'] > 0 else "—",
            f"{r['gross']:,.0f}"
        ]
        ws.append(row)
        
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

    ws.append([])
    
    # Formatlashni alohida bajaramiz
    total_gross_fmt = f"{total_gross:,.0f}".replace(',', ' ')
    tax_fmt = f"{tax:,.0f}".replace(',', ' ')
    net_fmt = f"{net:,.0f}".replace(',', ' ')
    
    summary_rows = [
        ['', '', '', '', '', 'JAMI (soliqsiz):', total_gross_fmt],
        ['', '', '', '', '', 'Soliq (7.5%):', tax_fmt],
        ['', '', '', '', '', 'QO\'LGA TEGADI:', net_fmt]
    ]
    
    for s_row in summary_rows:
        ws.append(s_row)
        for cell in ws[ws.max_row]:
            cell.border = border
            if s_row[5] in ['JAMI (soliqsiz):', 'Soliq (7.5%):', 'QO\'LGA TEGADI:']:
                cell.font = Font(bold=True)
                if s_row[5] == 'QO\'LGA TEGADI:':
                    ws.cell(row=ws.max_row, column=7).font = Font(bold=True, color="006100")
                    ws.cell(row=ws.max_row, column=7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    column_widths = [20, 12, 10, 10, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@dp.callback_query(F.data == "admin_monthly_report")
async def admin_monthly_report_start(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    now = datetime.now(UZB_TZ)
    start_date = datetime(2026, 3, 1, tzinfo=UZB_TZ) 
    
    builder = InlineKeyboardBuilder()
    months_uz = {1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun", 
                 7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"}
    
    temp_date = start_date
    while temp_date <= now:
        month_name = months_uz[temp_date.month]
        btn_text = f"📅 {month_name} {temp_date.year}"
        builder.row(InlineKeyboardButton(text=btn_text, callback_data=f"gen_month_{temp_date.year}_{temp_date.month}"))
        
        if temp_date.month == 12:
            temp_date = datetime(temp_date.year + 1, 1, 1, tzinfo=UZB_TZ)
        else:
            temp_date = datetime(temp_date.year, temp_date.month + 1, 1, tzinfo=UZB_TZ)
            
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    await callback.message.edit_text("Qaysi oy uchun hisobot kerak?", reply_markup=builder.as_markup())
    await callback.answer()

async def create_monthly_grouped_pdf(year: int, month: int) -> io.BytesIO:
    import calendar
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), topMargin=20)
    elements = []
    styles = getSampleStyleSheet()
    
    months_uz = {1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun", 
                 7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"}
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=20,
        alignment=1,
        spaceAfter=20,
        fontName=FONT_NAME_BOLD
    )
    day_style = ParagraphStyle(
        'DayHeader',
        parent=styles['Heading2'],
        fontSize=14,
        backColor=colors.lightgrey,
        spaceBefore=15,
        spaceAfter=5,
        fontName=FONT_NAME_BOLD
    )
    
    elements.append(Paragraph(f"{months_uz[month]} {year} - Oylik Davomat Hisoboti", title_style))
    _, last_day = calendar.monthrange(year, month)
    
    for day in range(1, last_day + 1):
        target_date = f"{year}-{month:02d}-{day:02d}"
        day_atts = [att for att in daily_attendance_log if att[2] == target_date]
        if not day_atts: continue 
        
        date_obj = d_date(year, month, day)
        weekday = WEEKDAYS_UZ[date_obj.weekday()]
        elements.append(Paragraph(f"📅 {target_date} ({weekday})", day_style))
        
        data = [['Vaqt', 'O\'qituvchi', 'Mutaxassislik', 'Filial', 'Holat', 'Kechikish']]
        for uid, branch, date_str, att_time in sorted(day_atts, key=lambda x: x[3]):
            teacher_name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            status, late_val = "Noma'lum", "-"
            if uid in user_schedules:
                for s_id in user_schedules[uid]:
                    s = schedules.get(s_id)
                    if s and s['branch'] == branch and weekday in s['days']:
                        les_time = s['days'][weekday]
                        ontime, mins = calculate_lateness(att_time, les_time)
                        status = "Vaqtida" if ontime else "Kechikkan"
                        late_val = "0" if ontime else f"{mins} min"
            data.append([att_time, teacher_name, specialty, branch, status, late_val])
            
        table = Table(data, colWidths=[0.8*inch, 2.2*inch, 1.3*inch, 2*inch, 1*inch, 1*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), FONT_NAME_BOLD),
            ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9F9')])
        ]))
        elements.append(table)
        elements.append(Spacer(1, 10))
        
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

@dp.callback_query(F.data.startswith("gen_month_"))
async def process_month_gen(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    _, _, year, month = callback.data.split("_")
    await callback.message.answer(f"⏳ {month}/{year} uchun hisobot tayyorlanmoqda...")
    pdf = await create_monthly_grouped_pdf(int(year), int(month))
    await callback.message.answer_document(
        types.BufferedInputFile(pdf.read(), filename=f"hisobot_{year}_{month}.pdf"),
        caption=f"📊 {month}/{year} oylik davomat hisoboti"
    )
    await callback.answer()

# --- PROFESSIONAL EXCEL HISOBOT (FILIALLAR BO'YICHA GURUHLANGAN) ---
async def create_monthly_excel(year: int, month: int) -> io.BytesIO:
    import calendar
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(border_style="thin", color="000000")
    all_border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    main_header_fill = PatternFill(start_color="92D050", fill_type="solid") # Yashil sarlavha
    user_header_fill = PatternFill(start_color="D9D9D9", fill_type="solid") # Kulrang o'qituvchi nomi
    branch_header_fill = PatternFill(start_color="FDE9D9", fill_type="solid") # Och jigarrang filial nomi
    table_header_fill = PatternFill(start_color="2E86AB", fill_type="solid") # Ko'k jadval header
    
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    
    months_uz = {1: "YANVAR", 2: "FEVRAL", 3: "MART", 4: "APREL", 5: "MAY", 6: "IYUN", 
                 7: "IYUL", 8: "AVGUST", 9: "SENTABR", 10: "OKTABR", 11: "NOYABR", 12: "DEKABR"}
    
    specs = ["IT", "Koreys tili", "Ofis xodimi"]
    _, last_day = calendar.monthrange(year, month)

    for spec in specs:
        ws = wb.create_sheet(title=spec)
        
        # 1. Asosiy sarlavha
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        main_title = ws.cell(row=1, column=1)
        main_title.value = f"{spec.upper()} OQITUVCHILARI {months_uz[month]} OYI XISOBOTI"
        main_title.fill = main_header_fill
        main_title.font = Font(size=14, bold=True)
        main_title.alignment = Alignment(horizontal="center")
        for col in range(1, 10):
            ws.cell(row=1, column=col).border = all_border
        
        current_row = 3
        teachers = [uid for uid, s in user_specialty.items() if s == spec]
        
        if not teachers:
            continue

        for uid in sorted(teachers, key=lambda x: user_names.get(x, "")):
            # 2. O'qituvchi nomi (Kulrang)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            user_title = ws.cell(row=current_row, column=1)
            user_title.value = f"👤 {user_names.get(uid, '').upper()} {months_uz[month]} OYLIK XISOBOT"
            user_title.fill = user_header_fill
            user_title.font = bold_font
            user_title.alignment = Alignment(horizontal="center")
            for col in range(1, 10):
                ws.cell(row=current_row, column=col).border = all_border
            current_row += 1

            # 3. Ushbu o'qituvchi ishlaydigan filiallarni aniqlash
            teacher_branches = sorted(list(set([s['branch'] for sid, s in schedules.items() if s['user_id'] == uid])))
            
            for branch in teacher_branches:
                # 4. Filial sarlavhasi (Och jigarrang)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
                branch_title = ws.cell(row=current_row, column=1)
                branch_title.value = f"🏢 Filial: {branch}"
                branch_title.fill = branch_header_fill
                branch_title.font = Font(italic=True, bold=True)
                branch_title.alignment = Alignment(horizontal="center")
                for col in range(1, 10):
                    ws.cell(row=current_row, column=col).border = all_border
                current_row += 1

                # 5. Jadval headeri
                headers = ['№', 'Sana', 'Hafta kuni', 'O\'qituvchi', 'Filial', 'Dars vaqti', 'Kelgan vaqti', 'Holat', 'Kechikish']
                for col, text in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = text
                    cell.fill = table_header_fill
                    cell.font = white_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = all_border
                current_row += 1

                # 6. Ma'lumotlarni yozish
                num = 1
                for d in range(1, last_day + 1):
                    target_date = f"{year}-{month:02d}-{d:02d}"
                    d_obj = d_date(year, month, d)
                    weekday = WEEKDAYS_UZ[d_obj.weekday()]
                    
                    # Jadvaldagi darsni qidirish
                    for sid, sdata in schedules.items():
                        if sdata['user_id'] == uid and sdata['branch'] == branch and weekday in sdata['days']:
                            sch_time = sdata['days'][weekday]
                            att = next((a for a in daily_attendance_log if a[0] == uid and a[1] == branch and a[2] == target_date), None)
                            
                            if att:
                                ontime, mins = calculate_lateness(att[3], sch_time)
                                status_text = "Vaqtida" if ontime else "Kechikkan"
                                att_time = att[3]
                                late_m = 0 if ontime else mins
                            else:
                                status_text = "KELMAGAN"
                                att_time = "—"
                                late_m = "—"

                            row_vals = [num, target_date, weekday, user_names.get(uid), branch, sch_time, att_time, status_text, late_m]
                            for col_idx, val in enumerate(row_vals, 1):
                                cell = ws.cell(row=current_row, column=col_idx)
                                cell.value = val
                                cell.border = all_border
                                cell.alignment = Alignment(horizontal="center")
                                
                                # Holat ustuniga rang berish
                                if col_idx == 8:
                                    if status_text == "Kechikkan":
                                        cell.font = Font(color="FF0000", bold=True)
                                    elif status_text == "Vaqtida":
                                        cell.font = Font(color="008000", bold=True)
                                    elif status_text == "KELMAGAN":
                                        cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")
                            
                            num += 1
                            current_row += 1
                current_row += 1 # Filiallar orasida kichik masofa
            current_row += 2 # O'qituvchilar orasida katta masofa

        # Ustun kengligi
        for i in range(1, 10):
            column_letter = get_column_letter(i)
            ws.column_dimensions[column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@dp.callback_query(F.data == "admin_excel_menu")
async def admin_excel_report_start(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    now = datetime.now(UZB_TZ)
    start_date = datetime(2026, 3, 1, tzinfo=UZB_TZ) 
    
    builder = InlineKeyboardBuilder()
    months_uz = {1: "Yanvar", 2: "Fevral", 3: "Mart", 4: "Aprel", 5: "May", 6: "Iyun", 
                 7: "Iyul", 8: "Avgust", 9: "Sentabr", 10: "Oktabr", 11: "Noyabr", 12: "Dekabr"}
    
    temp_date = start_date
    while temp_date <= now:
        month_name = months_uz[temp_date.month]
        btn_text = f"📊 Excel: {month_name} {temp_date.year}"
        builder.row(InlineKeyboardButton(text=btn_text, callback_data=f"get_excel_{temp_date.year}_{temp_date.month}"))
        
        if temp_date.month == 12:
            temp_date = datetime(temp_date.year + 1, 1, 1, tzinfo=UZB_TZ)
        else:
            temp_date = datetime(temp_date.year, temp_date.month + 1, 1, tzinfo=UZB_TZ)
            
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    await callback.message.edit_text("Qaysi oy uchun Excel hisobot kerak?", reply_markup=builder.as_markup())
    await callback.answer()

@dp.callback_query(F.data.startswith("get_excel_"))
async def process_excel_download(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    _, _, year, month = callback.data.split("_")
    await callback.message.answer(f"⏳ {month}-{year} uchun Excel tayyorlanmoqda...")
    
    try:
        excel_buf = await create_monthly_excel(int(year), int(month))
        
        filename = f"Davomat_{year}_{month}.xlsx"
        await callback.message.answer_document(
            types.BufferedInputFile(excel_buf.read(), filename=filename),
            caption=f"📈 {month}-{year} oyi uchun buxgalteriya hisoboti tayyor."
        )
    except Exception as e:
        logging.error(f"Excel yaratishda xatolik: {e}")
        await callback.message.answer(f"❌ Excel yaratishda xatolik: {e}")
    
    await callback.answer()

@dp.callback_query(F.data == "admin_pdf_menu")
async def admin_pdf_menu(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text="📊 Umumiy statistika", callback_data="pdf_general"),
        InlineKeyboardButton(text="🏆 Filiallar reytingi", callback_data="pdf_branches")
    )
    builder.row(
        InlineKeyboardButton(text="👥 O'qituvchilar reytingi", callback_data="pdf_teachers"),
        InlineKeyboardButton(text="📅 Oylik hisobot", callback_data="pdf_monthly")
    )
    builder.row(
        InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
    )
    
    await callback.message.edit_text(
        "📊 Statistika hisobotlari\n\nKerakli hisobot turini tanlang:",
        reply_markup=builder.as_markup()
    )
    await callback.answer()

async def create_general_stats_pdf() -> io.BytesIO:
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,
        spaceAfter=30,
        fontName=FONT_NAME_BOLD
    )
    elements.append(Paragraph("📊 Umumiy statistika", title_style))
    elements.append(Paragraph(f"Hisobot yaratilgan sana: {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    now_uzb = datetime.now(UZB_TZ)
    today = now_uzb.strftime("%Y-%m-%d")
    current_month = now_uzb.strftime("%Y-%m")
    
    total_users = len(user_ids)
    active_users = len([uid for uid in user_ids if user_status.get(uid) == 'active'])
    blocked_users = len([uid for uid in user_ids if user_status.get(uid) == 'blocked'])
    total_attendances = len(daily_attendance_log)
    today_attendances = len([k for k in daily_attendance_log if k[2] == today])
    monthly_attendances = len([k for k in daily_attendance_log if k[2].startswith(current_month)])
    
    it_teachers = len([uid for uid in user_ids if user_specialty.get(uid) == 'IT'])
    korean_teachers = len([uid for uid in user_ids if user_specialty.get(uid) == 'Koreys tili'])
    office_workers = len([uid for uid in user_ids if user_specialty.get(uid) == 'Ofis xodimi'])
    
    data = [
        ['Ko\'rsatkich', 'Qiymat'],
        ['Jami foydalanuvchilar', str(total_users)],
        ['Faol foydalanuvchilar', str(active_users)],
        ['Bloklanganlar', str(blocked_users)],
        ['💻 IT o\'qituvchilar', str(it_teachers)],
        ['🇰🇷 Koreys tili o\'qituvchilar', str(korean_teachers)],
        ['🏢 Ofis xodimlari', str(office_workers)],
        ['Jami davomatlar', str(total_attendances)],
        ['Bugungi davomatlar', str(today_attendances)],
        ['Shu oydagi davomatlar', str(monthly_attendances)],
    ]
    
    table = Table(data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_NAME_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 12)
    ]))
    elements.append(table)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

async def create_branches_stats_pdf() -> io.BytesIO:
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,
        spaceAfter=30,
        fontName=FONT_NAME_BOLD
    )
    elements.append(Paragraph("🏆 Filiallar reytingi", title_style))
    elements.append(Paragraph(f"Hisobot yaratilgan sana: {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    branch_stats = defaultdict(int)
    for (uid, branch, date, time) in daily_attendance_log:
        branch_stats[branch] += 1
    
    sorted_branches = sorted(branch_stats.items(), key=lambda x: x[1], reverse=True)
    
    data = [['№', 'Filial', 'Davomatlar soni']]
    for i, (branch, count) in enumerate(sorted_branches, 1):
        data.append([str(i), branch, str(count)])
    
    table = Table(data, colWidths=[0.5*inch, 3.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_NAME_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

async def create_teachers_stats_pdf() -> io.BytesIO:
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,
        spaceAfter=30,
        fontName=FONT_NAME_BOLD
    )
    elements.append(Paragraph("👥 O'qituvchilar reytingi", title_style))
    elements.append(Paragraph(f"Hisobot yaratilgan sana: {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    teacher_stats = defaultdict(int)
    teacher_info = {}
    
    for (uid, branch, date, time) in daily_attendance_log:
        teacher_stats[uid] += 1
        if uid not in teacher_info:
            teacher_info[uid] = {
                'name': user_names.get(uid, f"ID: {uid}"),
                'specialty': user_specialty.get(uid, '')
            }
    
    sorted_teachers = sorted(teacher_stats.items(), key=lambda x: x[1], reverse=True)
    
    data = [['№', 'O\'qituvchi', 'Mutaxassislik', 'Davomatlar']]
    for i, (uid, count) in enumerate(sorted_teachers[:50], 1):
        info = teacher_info.get(uid, {'name': f"ID: {uid}", 'specialty': ''})
        data.append([str(i), info['name'], info['specialty'], str(count)])
    
    table = Table(data, colWidths=[0.5*inch, 2.5*inch, 1.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_NAME_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

async def create_monthly_stats_pdf() -> io.BytesIO:
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    now_uzb = datetime.now(UZB_TZ)
    current_month = now_uzb.strftime("%Y-%m")
    month_name = now_uzb.strftime("%B %Y")
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,
        spaceAfter=30,
        fontName=FONT_NAME_BOLD
    )
    elements.append(Paragraph(f"📅 {month_name} oyi hisoboti", title_style))
    elements.append(Paragraph(f"Hisobot yaratilgan sana: {now_uzb.strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    monthly_stats = defaultdict(lambda: defaultdict(int))
    monthly_teacher_set = set()
    
    for (uid, branch, date, time) in daily_attendance_log:
        if date.startswith(current_month):
            monthly_stats[branch][uid] += 1
            monthly_teacher_set.add(uid)
    
    data = [['№', 'Filial', 'Jami davomatlar', 'O\'qituvchilar']]
    for i, (branch, users) in enumerate(sorted(monthly_stats.items(), key=lambda x: sum(x[1].values()), reverse=True), 1):
        total = sum(users.values())
        unique_teachers = len(users)
        data.append([str(i), branch, str(total), str(unique_teachers)])
    
    total_attendances = sum(sum(users.values()) for users in monthly_stats.values())
    total_teachers = len(monthly_teacher_set)
    total_branches = len(monthly_stats)
    
    data.append(['', 'UMUMIY:', str(total_attendances), str(total_teachers)])
    
    table = Table(data, colWidths=[0.5*inch, 3*inch, 1.5*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), FONT_NAME_BOLD),
        ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (-2, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (-2, -1), (-1, -1), FONT_NAME_BOLD),
    ]))
    elements.append(table)
    
    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

@dp.callback_query(F.data.startswith("pdf_"))
async def handle_pdf_reports(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    report_type = callback.data.replace("pdf_", "")
    
    await callback.message.edit_text("⏳ PDF hisobot tayyorlanmoqda, biroz kuting...")
    
    try:
        if report_type == "general":
            pdf_buffer = await create_general_stats_pdf()
            filename = f"umumiy_statistika_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.pdf"
            caption = "📊 Umumiy statistika hisoboti"
            
        elif report_type == "branches":
            pdf_buffer = await create_branches_stats_pdf()
            filename = f"filiallar_reytingi_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.pdf"
            caption = "🏆 Filiallar reytingi"
            
        elif report_type == "teachers":
            pdf_buffer = await create_teachers_stats_pdf()
            filename = f"oqituvchilar_reytingi_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.pdf"
            caption = "👥 Eng faol o'qituvchilar"
            
        elif report_type == "monthly":
            pdf_buffer = await create_monthly_stats_pdf()
            filename = f"oylik_hisobot_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.pdf"
            caption = f"📅 {datetime.now(UZB_TZ).strftime('%B %Y')} oyi hisoboti"
            
        else:
            await callback.message.edit_text("❌ Noto'g'ri so'rov")
            await callback.answer()
            return
        
        await callback.message.delete()
        await callback.message.answer_document(
            types.BufferedInputFile(pdf_buffer.getvalue(), filename=filename),
            caption=caption
        )
        
    except Exception as e:
        logging.error(f"PDF yaratishda xatolik: {e}")
        await callback.message.edit_text(f"❌ PDF yaratishda xatolik: {str(e)}")
    
    await callback.answer()

@dp.callback_query(F.data == "admin_stats_main")
async def admin_stats_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="📊 Umumiy statistika", callback_data="admin_stats_general"),
            InlineKeyboardButton(text="🏆 Filiallar reytingi", callback_data="admin_stats_branches")
        )
        builder.row(
            InlineKeyboardButton(text="👥 O'qituvchilar reytingi", callback_data="admin_stats_teachers"),
            InlineKeyboardButton(text="📅 Oylik hisobot", callback_data="admin_monthly")
        )
        builder.row(
            InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
        )
        
        await callback.message.edit_text(
            "📊 Statistika bo'limi\n\nKerakli statistikani tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_main error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data == "admin_stats_general")
async def admin_stats_general(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        now_uzb = datetime.now(UZB_TZ)
        today = now_uzb.strftime("%Y-%m-%d")
        current_month = now_uzb.strftime("%Y-%m")
        
        total_users = len(user_ids)
        active_users = len([uid for uid in user_ids if user_status.get(uid) == 'active'])
        blocked_users = len([uid for uid in user_ids if user_status.get(uid) == 'blocked'])
        total_attendances = len(daily_attendance_log)
        today_attendances = len([k for k in daily_attendance_log if k[2] == today])
        monthly_attendances = len([k for k in daily_attendance_log if k[2].startswith(current_month)])
        
        it_teachers = len([uid for uid in user_ids if user_specialty.get(uid) == 'IT'])
        korean_teachers = len([uid for uid in user_ids if user_specialty.get(uid) == 'Koreys tili'])
        office_workers = len([uid for uid in user_ids if user_specialty.get(uid) == 'Ofis xodimi'])
        
        branch_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            branch_stats[branch] += 1
        top_branch = max(branch_stats.items(), key=lambda x: x[1]) if branch_stats else ("Yo'q", 0)
        
        teacher_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            teacher_stats[uid] += 1
        top_teacher_id = max(teacher_stats.items(), key=lambda x: x[1]) if teacher_stats else (None, 0)
        top_teacher_name = user_names.get(top_teacher_id[0], "Noma'lum") if top_teacher_id[0] else "Yo'q"
        top_teacher_specialty = user_specialty.get(top_teacher_id[0], '')
        top_teacher_display = f"{top_teacher_name}[{top_teacher_specialty}]" if top_teacher_specialty else top_teacher_name
        
        text = f"""
📊 Umumiy statistika

👥 Foydalanuvchilar:
• Jami: {total_users}
• Faol: {active_users}
• Bloklangan: {blocked_users}
• 💻 IT: {it_teachers}
• 🇰🇷 Koreys tili: {korean_teachers}
• 🏢 Ofis xodimlari: {office_workers}

📋 Davomatlar:
• Jami: {total_attendances}
• Bugun: {today_attendances}
• Shu oyda: {monthly_attendances}

🏆 Eng faol filial: {top_branch[0]} ({top_branch[1]} ta)

👑 Eng faol o'qituvchi: {top_teacher_display} ({top_teacher_id[1]} ta)

📅 Oxirgi yangilanish: {now_uzb.strftime('%Y-%m-%d %H:%M')}
"""
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            text,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_general error: {e}")
        await callback.message.edit_text("❌ Statistikani olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_stats_branches")
async def admin_stats_branches(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        branch_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            branch_stats[branch] += 1
        
        if not branch_stats:
            await callback.message.edit_text("📭 Hali davomat ma'lumotlari yo'q.")
            await callback.answer()
            return
        
        sorted_branches = sorted(branch_stats.items(), key=lambda x: x[1], reverse=True)
        
        text = "🏆 Filiallar reytingi\n\n"
        for i, (branch, count) in enumerate(sorted_branches, 1):
            medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            text += f"{medal} {branch}: {count} ta davomat\n"
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            text,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_branches error: {e}")
        await callback.message.edit_text("❌ Filiallar reytingini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_stats_teachers")
async def admin_stats_teachers(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        teacher_stats = defaultdict(int)
        for (uid, branch, date, time) in daily_attendance_log:
            teacher_stats[uid] += 1
        
        if not teacher_stats:
            await callback.message.edit_text("📭 Hali davomat ma'lumotlari yo'q.")
            await callback.answer()
            return
        
        sorted_teachers = sorted(teacher_stats.items(), key=lambda x: x[1], reverse=True)[:20]
        
        text = "👥 Eng faol o'qituvchilar\n\n"
        for i, (uid, count) in enumerate(sorted_teachers, 1):
            name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            specialty_display = f" [{specialty}]" if specialty else ""
            medal = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            text += f"{medal} {name}{specialty_display}: {count} ta davomat\n"
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            text,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_stats_teachers error: {e}")
        await callback.message.edit_text("❌ O'qituvchilar reytingini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_monthly")
async def admin_monthly(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        now_uzb = datetime.now(UZB_TZ)
        current_month = now_uzb.strftime("%Y-%m")
        month_name = now_uzb.strftime("%B %Y")
        
        monthly_stats = defaultdict(lambda: defaultdict(int))
        for (uid, branch, date, time) in daily_attendance_log:
            if date.startswith(current_month):
                monthly_stats[branch][uid] += 1
        
        if not monthly_stats:
            await callback.message.edit_text("📭 Shu oy uchun davomat ma'lumotlari yo'q.")
            await callback.answer()
            return
        
        report = f"📊 {month_name} oyi uchun hisobot\n\n"
        
        for branch, users in monthly_stats.items():
            total = sum(users.values())
            unique_users = len(users)
            report += f"🏢 {branch}\n"
            report += f"   Jami: {total} ta davomat\n"
            report += f"   O'qituvchilar: {unique_users} ta\n\n"
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_stats_main"))
        
        await callback.message.edit_text(
            report,
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_monthly error: {e}")
        await callback.message.edit_text("❌ Oylik hisobotni olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_users_main")
async def admin_users_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="✅ Foydalanuvchilar ro'yxati", callback_data="admin_users_active")
        )
        builder.row(
            InlineKeyboardButton(text="⛔ Bloklanganlar", callback_data="admin_users_blocked")
        )
        builder.row(
            InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
        )
        
        await callback.message.edit_text(
            "👥 Foydalanuvchilarni boshqarish\n\nKerakli amalni tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_users_main error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data.startswith("admin_user_info_"))
async def admin_user_info(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    logging.info(f"admin_user_info handler: {callback.data}")
    
    try:
        uid = int(callback.data.replace("admin_user_info_", ""))
        logging.info(f"admin_user_info called for uid: {uid}")
    except ValueError as e:
        logging.error(f"admin_user_info parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    name = user_names.get(uid, "Noma'lum")
    status = user_status.get(uid, 'active')
    lang = user_languages.get(uid, 'uz')
    specialty = user_specialty.get(uid, '')
    specialty_display = f" [{specialty}]" if specialty else ""
    
    user_attendances = len([k for k in daily_attendance_log if k[0] == uid])
    user_schedules_count = len(user_schedules.get(uid,[]))
    
    last_attendance = "Yo'q"
    user_logs = [k for k in daily_attendance_log if k[0] == uid]
    if user_logs:
        last = max(user_logs, key=lambda x: x[2])
        last_attendance = f"{last[2]} {last[3]} ({last[1]})"
    
    text = f"""
👤 Foydalanuvchi ma'lumoti

ID: `{uid}`
Ism: {name}{specialty_display}
Holat: {"✅ Faol" if status == 'active' else "⛔ Bloklangan"}
Til: {lang}

📊 Statistika:
• Jami davomatlar: {user_attendances}
• Dars jadvallari: {user_schedules_count}
• Oxirgi davomat: {last_attendance}
"""
    
    builder = InlineKeyboardBuilder()
    if status == 'active':
        builder.row(InlineKeyboardButton(text="⛔ Bloklash", callback_data=f"admin_user_block_{uid}"))
    else:
        builder.row(InlineKeyboardButton(text="✅ Faollashtirish", callback_data=f"admin_user_unblock_{uid}"))
    builder.row(
        InlineKeyboardButton(text="📊 Statistika", callback_data=f"admin_user_stats_{uid}"),
        InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"admin_user_delete_{uid}")
    )
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
    
    await callback.message.edit_text(
        text,
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(F.data.startswith("admin_user_block_"))
async def admin_user_block(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        uid = int(callback.data.replace("admin_user_block_", ""))
        logging.info(f"admin_user_block called for uid: {uid}")
    except ValueError as e:
        logging.error(f"admin_user_block parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    user_status[uid] = 'blocked'
    await db.update_user_status(uid, 'blocked')
    
    await callback.answer("✅ Foydalanuvchi bloklandi!")
    await admin_user_info(callback)

@dp.callback_query(F.data.startswith("admin_user_unblock_"))
async def admin_user_unblock(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        uid = int(callback.data.replace("admin_user_unblock_", ""))
        logging.info(f"admin_user_unblock called for uid: {uid}")
    except ValueError as e:
        logging.error(f"admin_user_unblock parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    user_status[uid] = 'active'
    await db.update_user_status(uid, 'active')
    
    await callback.answer("✅ Foydalanuvchi faollashtirildi!")
    await admin_user_info(callback)

@dp.callback_query(F.data.startswith("admin_user_delete_confirm_"))
async def admin_user_delete_confirm(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    logging.info(f"admin_user_delete_confirm handler: {callback.data}")
    
    try:
        uid = int(callback.data.replace("admin_user_delete_confirm_", ""))
        logging.info(f"admin_user_delete_confirm called for uid: {uid}")
        
        user_name = user_names.get(uid, "Noma'lum")
        user_spec = user_specialty.get(uid, "")
        spec_display = f" [{user_spec}]" if user_spec else ""
        
        await callback.message.edit_text(
            f"⏳ Foydalanuvchi o'chirilmoqda...\n\nID: `{uid}`\nIsm: {user_name}{spec_display}"
        )
        
    except ValueError as e:
        logging.error(f"admin_user_delete_confirm parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    
    try:
        async with db.pool.acquire() as conn:
            await conn.execute("DELETE FROM attendance WHERE user_id = $1", uid)
            await conn.execute("DELETE FROM schedules WHERE user_id = $1", uid)
            await conn.execute("DELETE FROM users WHERE user_id = $1", uid)
        
        if uid in user_ids:
            user_ids.remove(uid)
        
        user_names.pop(uid, None)
        user_specialty.pop(uid, None)
        user_status.pop(uid, None)
        user_languages.pop(uid, None)
        
        to_remove = [k for k in daily_attendance_log if k[0] == uid]
        for k in to_remove:
            daily_attendance_log.discard(k)
        
        keys_to_remove = [k for k in attendance_counter.keys() if k[0] == uid]
        for k in keys_to_remove:
            del attendance_counter[k]
        
        if uid in user_schedules:
            for schedule_id in user_schedules[uid]:
                schedules.pop(schedule_id, None)
            user_schedules.pop(uid, None)
        
        await callback.message.edit_text(
            f"✅ Foydalanuvchi muvaffaqiyatli o'chirildi!\n\n"
            f"ID: `{uid}`\n"
            f"Ism: {user_name}{spec_display}\n\n"
            f"Barcha ma'lumotlari bazadan tozalandi."
        )
        
        await callback.answer("✅ Foydalanuvchi o'chirildi!")
        
        await asyncio.sleep(2)
        
        active_users = [uid for uid in user_ids if user_status.get(uid) != 'blocked']
        if active_users:
            builder = InlineKeyboardBuilder()
            for uid in sorted(active_users)[:15]:
                name = user_names.get(uid, f"ID: {uid}")
                specialty = user_specialty.get(uid, '')
                spec_display = f" [{specialty}]" if specialty else ""
                
                if len(name) > 30:
                    name = name[:27] + "..."
                    
                builder.row(
                    InlineKeyboardButton(
                        text=f"👤 {name}{spec_display}", 
                        callback_data=f"admin_user_info_{uid}"
                    )
                )
            
            builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
            
            await callback.message.answer(
                "✅ Faol foydalanuvchilar ro'yxati:",
                reply_markup=builder.as_markup()
            )
        else:
            await callback.message.answer("📭 Faol foydalanuvchilar yo'q.")
        
    except Exception as e:
        logging.error(f"admin_user_delete_confirm error: {e}")
        traceback.print_exc()
        
        await callback.message.edit_text(
            f"❌ Xatolik yuz berdi: {str(e)}\n\n"
            f"Batafsil ma'lumot uchun loglarni tekshiring."
        )
        await callback.answer("Xatolik yuz berdi!")

@dp.callback_query(F.data.regexp(r"^admin_user_delete_\d+$"))
async def admin_user_delete(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    logging.info(f"admin_user_delete handler: {callback.data}")
    
    try:
        uid = int(callback.data.replace("admin_user_delete_", ""))
        logging.info(f"admin_user_delete called for uid: {uid}")
        
        user_name = user_names.get(uid, "Noma'lum")
        user_spec = user_specialty.get(uid, "")
        spec_display = f" [{user_spec}]" if user_spec else ""
        
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="✅ Ha, o'chirish", callback_data=f"admin_user_delete_confirm_{uid}"),
            InlineKeyboardButton(text="❌ Bekor qilish", callback_data=f"admin_user_info_{uid}")
        )
        
        await callback.message.edit_text(
            f"⚠️ Foydalanuvchini o'chirish\n\n"
            f"ID: `{uid}`\n"
            f"Ism: {user_name}{spec_display}\n\n"
            f"Bu foydalanuvchini butunlay o'chirmoqchimisiz?\n"
            f"Barcha ma'lumotlari (davomatlar, dars jadvallari) ham o'chib ketadi!",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
        
    except ValueError as e:
        logging.error(f"admin_user_delete parse error: {callback.data}, error: {e}")
        await callback.answer("Noto'g'ri format!")
        return
    except Exception as e:
        logging.error(f"admin_user_delete error: {e}")
        traceback.print_exc()
        await callback.answer("Xatolik yuz berdi!")

@dp.callback_query(F.data.startswith("admin_user_stats_"))
async def admin_user_stats(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        uid = int(callback.data.replace("admin_user_stats_", ""))
    except ValueError:
        await callback.answer("Noto'g'ri format!")
        return
    
    name = user_names.get(uid, "Noma'lum")
    
    branch_stats = defaultdict(int)
    month_stats = defaultdict(int)
    
    for (user_id, branch, date, time) in daily_attendance_log:
        if user_id == uid:
            branch_stats[branch] += 1
            month = date[:7]
            month_stats[month] += 1
    
    text = f"📊 {name} statistikasi\n\n"
    
    if branch_stats:
        text += "🏢 Filiallar bo'yicha:\n"
        for branch, count in sorted(branch_stats.items(), key=lambda x: x[1], reverse=True):
            text += f"• {branch}: {count} ta\n"
        text += "\n"
    
    if month_stats:
        text += "📅 Oylar bo'yicha:\n"
        for month, count in sorted(month_stats.items(), reverse=True):
            text += f"• {month}: {count} ta\n"
    else:
        text += "📭 Hali davomat yo'q"
    
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data=f"admin_user_info_{uid}"))
    
    await callback.message.edit_text(
        text,
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(F.data == "admin_users_active")
async def admin_users_active(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        active =[uid for uid in user_ids if user_status.get(uid) != 'blocked']
        
        if not active:
            await callback.message.edit_text("📭 Faol foydalanuvchilar yo'q.")
            await callback.answer()
            return
        
        builder = InlineKeyboardBuilder()
        for uid in sorted(active)[:20]:
            name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            specialty_display = f" [{specialty}]" if specialty else ""
            builder.row(
                InlineKeyboardButton(text=f"✅ {name}{specialty_display}", callback_data=f"admin_user_info_{uid}")
            )
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
        
        await callback.message.edit_text(
            "✅ Faol foydalanuvchilar ro'yxati:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_users_active error: {e}")
        await callback.message.edit_text("❌ Faol foydalanuvchilar ro'yxatini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_users_blocked")
async def admin_users_blocked(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        blocked =[uid for uid in user_ids if user_status.get(uid) == 'blocked']
        
        if not blocked:
            await callback.message.edit_text("📭 Bloklangan foydalanuvchilar yo'q.")
            await callback.answer()
            return
        
        builder = InlineKeyboardBuilder()
        for uid in blocked[:20]:
            name = user_names.get(uid, f"ID: {uid}")
            specialty = user_specialty.get(uid, '')
            specialty_display = f" [{specialty}]" if specialty else ""
            builder.row(
                InlineKeyboardButton(text=f"⛔ {name}{specialty_display}", callback_data=f"admin_user_info_{uid}")
            )
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_users_main"))
        
        await callback.message.edit_text(
            "⛔ Bloklangan foydalanuvchilar:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_users_blocked error: {e}")
        await callback.message.edit_text("❌ Bloklangan foydalanuvchilar ro'yxatini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_broadcast")
async def admin_broadcast_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    user_id = callback.from_user.id
    lang = user_languages.get(user_id, 'uz')
    
    builder = InlineKeyboardBuilder()
    builder.row(
        InlineKeyboardButton(text=TRANSLATIONS[lang]['specialty_it'], callback_data="broadcast_spec_IT"),
        InlineKeyboardButton(text=TRANSLATIONS[lang]['specialty_korean'], callback_data="broadcast_spec_Koreys tili")
    )
    builder.row(
        InlineKeyboardButton(text="🏢 Ofis xodimi", callback_data="broadcast_spec_Ofis xodimi"),
        InlineKeyboardButton(text=TRANSLATIONS[lang]['all_teachers'], callback_data="broadcast_spec_all")
    )
    builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back"))
    
    await state.set_state(Broadcast.selecting_specialty)
    await callback.message.edit_text(
        get_text(user_id, 'select_broadcast_specialty'),
        reply_markup=builder.as_markup()
    )
    await callback.answer()

@dp.callback_query(Broadcast.selecting_specialty, F.data.startswith("broadcast_spec_"))
async def admin_broadcast_specialty(callback: types.CallbackQuery, state: FSMContext):
    try:
        specialty = callback.data.replace("broadcast_spec_", "")
        if specialty == "all":
            specialty = None
        
        await state.update_data(specialty=specialty)
        await state.set_state(Broadcast.waiting_for_message)
        
        await callback.message.edit_text(
            "📢 Xabar yuborish\n\nYubormoqchi bo'lgan xabaringizni kiriting (matn, rasm, hujjat):"
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_broadcast_specialty error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(Broadcast.waiting_for_message)
async def admin_broadcast_message(message: types.Message, state: FSMContext):
    if not check_admin(message.chat.id):
        await state.clear()
        return
    
    try:
        await state.update_data(
            message_text=message.text or message.caption,
            message_type=message.content_type,
            message_data=message
        )
        
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="✅ Ha", callback_data="broadcast_confirm"),
            InlineKeyboardButton(text="❌ Yo'q", callback_data="broadcast_cancel")
        )
        
        data = await state.get_data()
        specialty = data.get('specialty')
        
        if specialty:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked' and user_specialty.get(uid) == specialty]
            specialty_text = f" ({specialty})"
        else:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked']
            specialty_text = " (barcha)"
        
        total_users = len(target_users)
        
        await state.set_state(Broadcast.waiting_for_confirm)
        await message.answer(
            f"📢 Xabar yuborishni tasdiqlang{specialty_text}\n\n"
            f"Xabar: {message.text or 'Rasm/hujjat'}\n"
            f"Qabul qiluvchilar: {total_users} ta foydalanuvchi\n\n"
            f"Yuborishni boshlaymizmi?",
            reply_markup=builder.as_markup()
        )
    except Exception as e:
        logging.error(f"admin_broadcast_message error: {e}")
        await message.answer("❌ Xatolik yuz berdi")
        await state.clear()

@dp.callback_query(Broadcast.waiting_for_confirm, F.data == "broadcast_confirm")
async def admin_broadcast_confirm(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        data = await state.get_data()
        specialty = data.get('specialty')
        
        if specialty:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked' and user_specialty.get(uid) == specialty]
        else:
            target_users =[uid for uid in user_ids if user_status.get(uid) != 'blocked']
        
        await callback.message.edit_text("⏳ Xabarlar yuborilmoqda...")
        
        sent_count = 0
        failed_count = 0
        
        for user_id in target_users:
            try:
                msg_data = data['message_data']
                if data['message_type'] == 'text':
                    await bot.send_message(user_id, msg_data.text)
                elif data['message_type'] == 'photo':
                    await bot.send_photo(user_id, msg_data.photo[-1].file_id, caption=msg_data.caption)
                elif data['message_type'] == 'document':
                    await bot.send_document(user_id, msg_data.document.file_id, caption=msg_data.caption)
                sent_count += 1
                await asyncio.sleep(0.05)
            except Exception as e:
                failed_count += 1
                logging.error(f"Broadcast error for user {user_id}: {e}")
        
        broadcast_history.append({
            'text': data.get('message_text', ''),
            'date': datetime.now(UZB_TZ).strftime("%Y-%m-%d %H:%M:%S"),
            'sent_count': sent_count,
            'failed_count': failed_count,
            'specialty': specialty
        })
        
        await db.save_broadcast(data.get('message_text', ''), sent_count, failed_count, specialty)
        
        specialty_text = f" ({specialty})" if specialty else " (barcha)"
        
        await callback.message.edit_text(
            f"✅ Xabar yuborildi{specialty_text}!\n\n"
            f"✓ Yuborildi: {sent_count}\n"
            f"✗ Xatolik: {failed_count}"
        )
        
        await state.clear()
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_broadcast_confirm error: {e}")
        await callback.message.edit_text("❌ Xabar yuborishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(Broadcast.waiting_for_confirm, F.data == "broadcast_cancel")
async def admin_broadcast_cancel(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        await state.clear()
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Admin panel", callback_data="admin_back"))
        
        await callback.message.edit_text(
            "❌ Xabar yuborish bekor qilindi.",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_broadcast_cancel error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data == "admin_schedules_main")
async def admin_schedules_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="➕ O'qituvchiga jadval qo'shish", callback_data="admin_add_schedule")
        )
        builder.row(
            InlineKeyboardButton(text="📋 Faol dars jadvallari", callback_data="admin_active_schedules")
        )
        builder.row(
            InlineKeyboardButton(text="📋 Dars jadvali PDF", callback_data="admin_schedules_pdf")
        )
        builder.row(
            InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
        )
        
        await callback.message.edit_text(
            "📅 Dars jadvallarini boshqarish",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_schedules_main error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data == "admin_schedules_pdf")
async def admin_schedules_pdf(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    await callback.message.edit_text("⏳ Dars jadvallari PDF tayyorlanmoqda, biroz kuting...")
    
    try:
        pdf_buffer = await create_all_schedules_pdf()
        
        filename = f"barcha_dars_jadvallari_{datetime.now(UZB_TZ).strftime('%Y%m%d_%H%M')}.pdf"
        
        await callback.message.delete()
        await callback.message.answer_document(
            types.BufferedInputFile(pdf_buffer.getvalue(), filename=filename),
            caption="📋 Barcha o'qituvchilarning dars jadvallari"
        )
        
    except Exception as e:
        logging.error(f"admin_schedules_pdf error: {e}")
        await callback.message.edit_text(f"❌ PDF yaratishda xatolik: {str(e)}")
    
    await callback.answer()

async def create_all_schedules_pdf() -> io.BytesIO:
    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer, 
        pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
    )
    elements = []
    styles = getSampleStyleSheet()

    main_title_style = ParagraphStyle(
        'MainTitle', parent=styles['Heading1'], fontSize=26, alignment=1, 
        spaceAfter=10, textColor=colors.HexColor('#1A237E'), fontName=FONT_NAME_BOLD
    )
    
    section_style = ParagraphStyle(
        'SectionHeader', parent=styles['Heading2'], fontSize=18, alignment=1,
        textColor=colors.whitesmoke, backColor=colors.HexColor('#283593'),
        spaceBefore=10, spaceAfter=15, borderPadding=10, borderRadius=5,
        fontName=FONT_NAME_BOLD
    )

    teacher_name_style = ParagraphStyle(
        'TeacherName', parent=styles['Heading3'], fontSize=14, 
        textColor=colors.HexColor('#B71C1C'), spaceBefore=12, spaceAfter=5,
        fontName=FONT_NAME_BOLD
    )

    branch_info_style = ParagraphStyle(
        'BranchInfo', parent=styles['Normal'], fontSize=11, 
        textColor=colors.HexColor('#0D47A1'), fontName=FONT_NAME_BOLD
    )

    normal_style = ParagraphStyle(
        'Normal', parent=styles['Normal'], fontName=FONT_NAME
    )

    elements.append(Paragraph("HANCOM ACADEMY", main_title_style))
    elements.append(Paragraph(
        "O'QITUVCHILAR VA XODIMLARNING UMUMIY DARS JADVALI",
        ParagraphStyle('Sub', parent=normal_style, alignment=1, fontSize=12, spaceAfter=20, fontName=FONT_NAME)
    ))

    grouped_data = defaultdict(lambda: defaultdict(list))
    
    for s_id, s_data in schedules.items():
        uid = s_data['user_id']
        spec = user_specialty.get(uid, "Boshqa")
        t_name = user_names.get(uid, f"ID: {uid}")
        grouped_data[spec][t_name].append(s_data)

    spec_order = ["IT", "Koreys tili", "Ofis xodimi"]
    
    for spec in spec_order:
        if spec not in grouped_data:
            continue
        
        elements.append(Paragraph(f"{spec.upper()} BO'LIMI JADVALLARI", section_style))
        
        sorted_teachers = sorted(grouped_data[spec].keys())
        
        for t_name in sorted_teachers:
            teacher_block = []
            
            teacher_block.append(Paragraph(f"👤 O'qituvchi: {t_name}", teacher_name_style))
            
            teacher_scheds = grouped_data[spec][t_name]
            
            for s_data in teacher_scheds:
                branch = s_data['branch']
                l_type = s_data.get('lesson_type', 'Dars')
                
                teacher_block.append(Paragraph(f"📍 Filial: {branch} ({l_type})", branch_info_style))
                
                days = sort_weekdays(s_data['days'])
                table_data = [['Hafta kuni', 'Dars boshlanish vaqti']]
                
                for d_name, d_time in days.items():
                    table_data.append([d_name, d_time])
                
                t = Table(table_data, colWidths=[2.5*inch, 2.5*inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E3F2FD')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1565C0')),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), FONT_NAME_BOLD),
                    ('FONTNAME', (0, 1), (-1, -1), FONT_NAME),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                    ('TOPPADDING', (0, 0), (-1, -1), 5),
                ]))
                
                teacher_block.append(t)
                teacher_block.append(Spacer(1, 10))
            
            elements.append(KeepTogether(teacher_block))
            elements.append(Spacer(1, 5))
            elements.append(Paragraph("<hr/>", normal_style))
        
        elements.append(PageBreak())

    footer_text = f"Hisobot yaratildi: {datetime.now(UZB_TZ).strftime('%d.%m.%Y %H:%M')} | Hancom Academy Management System"
    elements.append(Paragraph(
        footer_text,
        ParagraphStyle('Footer', parent=normal_style, fontSize=8, alignment=2, textColor=colors.grey, fontName=FONT_NAME)
    ))

    doc.build(elements)
    pdf_buffer.seek(0)
    return pdf_buffer

@dp.callback_query(F.data == "admin_active_schedules")
async def admin_active_schedules(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        if not schedules:
            await callback.message.edit_text("📭 Hali dars jadvallari mavjud emas.")
            await callback.answer()
            return
        
        for schedule_id, schedule in schedules.items():
            teacher_name = user_names.get(schedule['user_id'], f"ID: {schedule['user_id']}")
            teacher_specialty = user_specialty.get(schedule['user_id'], '')
            specialty_display = f" [{teacher_specialty}]" if teacher_specialty else ""
            branch = schedule['branch']
            lesson_type = schedule.get('lesson_type', 'Dars')
            days_times = ""
            for day, time in sorted(schedule['days'].items(), key=lambda x: WEEKDAY_ORDER.get(x[0], 0)):
                days_times += f"• {day}: {time}\n"
            
            text = f"{teacher_name}{specialty_display}\n🏢 {branch}\n📚 {lesson_type}\n{days_times}"
            
            builder = InlineKeyboardBuilder()
            builder.row(
                InlineKeyboardButton(text="✏️ O'zgartirish", callback_data=f"admin_edit_schedule_{schedule_id}"),
                InlineKeyboardButton(text="🗑 O'chirish", callback_data=f"admin_delete_schedule_{schedule_id}")
            )
            
            await callback.message.answer(
                text,
                reply_markup=builder.as_markup()
            )
        
        await callback.message.delete()
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_active_schedules error: {e}")
        await callback.message.edit_text("❌ Dars jadvallarini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data.startswith("admin_delete_schedule_"))
async def admin_delete_schedule(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        schedule_id = callback.data.replace("admin_delete_schedule_", "")
        
        if schedule_id in schedules:
            schedule = schedules[schedule_id]
            teacher_id = schedule['user_id']
            
            await db.delete_schedule(schedule_id)
            
            try:
                lang = user_languages.get(teacher_id, 'uz')
                await bot.send_message(
                    teacher_id,
                    get_text(teacher_id, 'schedule_deleted_notify')
                )
            except Exception as e:
                logging.error(f"Failed to notify teacher {teacher_id}: {e}")
            
            del schedules[schedule_id]
            if teacher_id in user_schedules and schedule_id in user_schedules[teacher_id]:
                user_schedules[teacher_id].remove(schedule_id)
            
            await callback.message.edit_text("✅ Dars jadvali o'chirildi!")
        else:
            await callback.message.edit_text("❌ Jadval topilmadi!")
        
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_delete_schedule error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data.startswith("admin_edit_schedule_"))
async def admin_edit_schedule_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        schedule_id = callback.data.replace("admin_edit_schedule_", "")
        schedule = schedules.get(schedule_id)
        
        if not schedule:
            await callback.message.edit_text("❌ Jadval topilmadi!")
            await callback.answer()
            return
        
        await state.update_data(edit_schedule_id=schedule_id)
        await state.update_data(original_schedule=schedule)
        
        builder = InlineKeyboardBuilder()
        for location in LOCATIONS:
            builder.row(
                InlineKeyboardButton(text=location['name'], callback_data=f"edit_branch_{location['name']}")
            )
        
        await state.set_state(AdminEditSchedule.editing_branch)
        await callback.message.edit_text(
            get_text(callback.from_user.id, 'select_new_branch'),
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_edit_schedule_start error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(AdminEditSchedule.editing_branch, F.data.startswith("edit_branch_"))
async def admin_edit_schedule_branch(callback: types.CallbackQuery, state: FSMContext):
    try:
        branch = callback.data.replace("edit_branch_", "")
        await state.update_data(edit_branch=branch)
        
        user_id = callback.from_user.id
        lang = user_languages.get(user_id, 'uz')
        lesson_types = LESSON_TYPES.get(lang, LESSON_TYPES['uz'])
        
        builder = InlineKeyboardBuilder()
        for lesson in lesson_types:
            builder.row(
                InlineKeyboardButton(text=lesson, callback_data=f"edit_lesson_{lesson}")
            )
        
        await state.set_state(AdminEditSchedule.editing_lesson_type)
        await callback.message.edit_text(
            get_text(user_id, 'select_new_lesson_type'),
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_edit_schedule_branch error: {e}")
        await callback.answer("❌ Xatolik yuz berdi")

@dp.callback_query(AdminEditSchedule.editing_lesson_type, F.data.startswith("edit_lesson_"))
async def admin_edit_schedule_lesson(callback: types.CallbackQuery, state: FSMContext):
    try:
        lesson_type = callback.data.replace("edit_lesson_", "")
        await state.update_data(edit_lesson_type=lesson_type)
        
        user_id = callback.from_user.id
        lang = user_languages.get(user_id, 'uz')
        weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
        
        builder = InlineKeyboardBuilder()
        for i, day in enumerate(weekdays):
            builder.row(
                InlineKeyboardButton(text=f"⬜ {day}", callback_data=f"edit_weekday_{i}")
            )
        builder.row(
            InlineKeyboardButton(text="➡️ Keyingisi", callback_data="edit_weekdays_next")
        )
        
        await state.update_data(edit_selected_days={})
        await state.set_state(AdminEditSchedule.editing_weekdays)
        await callback.message.edit_text(
            get_text(user_id, 'select_new_weekdays'),
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_edit_schedule_lesson error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(AdminEditSchedule.editing_weekdays, F.data.startswith("edit_weekday_"))
async def admin_edit_schedule_weekday_select(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        selected_days = data.get('edit_selected_days', {})
        day_index = int(callback.data.replace("edit_weekday_", ""))
        
        user_id = callback.from_user.id
        lang = user_languages.get(user_id, 'uz')
        weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
        
        if day_index in selected_days:
            del selected_days[day_index]
        else:
            selected_days[day_index] = None
        
        await state.update_data(edit_selected_days=selected_days)
        
        builder = InlineKeyboardBuilder()
        for i, day in enumerate(weekdays):
            if i in selected_days:
                builder.row(
                    InlineKeyboardButton(text=f"✅ {day}", callback_data=f"edit_weekday_{i}")
                )
            else:
                builder.row(
                    InlineKeyboardButton(text=f"⬜ {day}", callback_data=f"edit_weekday_{i}")
                )
        builder.row(
            InlineKeyboardButton(text="➡️ Keyingisi", callback_data="edit_weekdays_next")
        )
        
        await callback.message.edit_reply_markup(reply_markup=builder.as_markup())
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_edit_schedule_weekday_select error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(AdminEditSchedule.editing_weekdays, F.data == "edit_weekdays_next")
async def admin_edit_schedule_weekdays_next(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        selected_days = data.get('edit_selected_days', {})
        
        if not selected_days:
            await callback.answer("Hech bo'lmaganda 1 kun tanlang!", show_alert=True)
            return
        
        days_without_time = [day for day in selected_days if selected_days[day] is None]
        
        if days_without_time:
            await state.update_data(edit_current_day=days_without_time[0])
            await state.set_state(AdminEditSchedule.editing_time)
            
            user_id = callback.from_user.id
            lang = user_languages.get(user_id, 'uz')
            weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
            day_name = weekdays[days_without_time[0]]
            
            await callback.message.edit_text(
                get_text(user_id, 'enter_new_time', weekday=day_name)
            )
        else:
            await admin_save_edited_schedule(callback.message, state)
        
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_edit_schedule_weekdays_next error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(AdminEditSchedule.editing_time)
async def admin_edit_schedule_enter_time(message: types.Message, state: FSMContext):
    try:
        time_str = message.text.strip()
        hours, minutes = map(int, time_str.split(':'))
        if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
            raise ValueError
        formatted_time = f"{hours:02d}:{minutes:02d}"
    except:
        await message.answer("❌ Noto'g'ri format! Iltimos, HH:MM formatida kiriting (masalan: 09:00)")
        return
    
    data = await state.get_data()
    selected_days = data.get('edit_selected_days', {})
    current_day = data.get('edit_current_day')
    
    selected_days[current_day] = formatted_time
    await state.update_data(edit_selected_days=selected_days)
    
    days_without_time =[day for day in selected_days if selected_days[day] is None]
    
    if days_without_time:
        await state.update_data(edit_current_day=days_without_time[0])
        
        user_id = message.from_user.id
        lang = user_languages.get(user_id, 'uz')
        weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
        day_name = weekdays[days_without_time[0]]
        
        await message.answer(
            get_text(user_id, 'enter_new_time', weekday=day_name)
        )
    else:
        await admin_save_edited_schedule(message, state)

async def admin_save_edited_schedule(message: types.Message, state: FSMContext):
    data = await state.get_data()
    old_s_id = data.get('edit_schedule_id')
    old_schedule = schedules.get(old_s_id)
    teacher_id = old_schedule['user_id']
    new_branch = data.get('edit_branch')
    new_lesson_type = data.get('edit_lesson_type')
    new_selected_days_raw = data.get('edit_selected_days', {})
    
    new_days = {WEEKDAYS_UZ[idx]: time for idx, time in new_selected_days_raw.items()}

    await db.delete_schedule(old_s_id)
    schedules.pop(old_s_id, None)
    if old_s_id in user_schedules[teacher_id]:
        user_schedules[teacher_id].remove(old_s_id)

    new_s_id = f"sch_{teacher_id}_{datetime.now().timestamp()}"
    schedules[new_s_id] = {'user_id': teacher_id, 'branch': new_branch, 'lesson_type': new_lesson_type, 'days': new_days}
    user_schedules[teacher_id].append(new_s_id)
    await db.save_schedule(new_s_id, teacher_id, new_branch, new_lesson_type, new_days)

    old_times = ", ".join([f"{k}:{v}" for k, v in old_schedule['days'].items()])
    new_times = ", ".join([f"{k}:{v}" for k, v in new_days.items()])
    
    msg = (f"📢 DIQQAT: Dars jadvalingiz o'zgardi!\n\n"
           f"🏢 Filial: {new_branch}\n"
           f"❌ Eski vaqtlar: {old_times}\n"
           f"✅ Yangi vaqtlar: {new_times}\n\n"
           f"Yangi jadval PDF ko'rinishida quyida biriktirildi.")
    
    try:
        await bot.send_message(teacher_id, msg)
        pdf = await create_schedule_pdf(teacher_id)
        clean_name = clean_pdf_text(user_names.get(teacher_id, 'user'))
        await bot.send_document(
            teacher_id, 
            types.BufferedInputFile(pdf.read(), filename=f"Dars_Jadvali_{clean_name}.pdf"),
            caption="📅 Yangi dars jadvalingiz (Landscape formatda)"
        )
    except Exception as e:
        logging.error(f"Notify error: {e}")

    await message.answer("✅ Jadval yangilandi, eskisi o'chirildi va o'qituvchiga xabar yuborildi.")
    await state.clear()

@dp.callback_query(F.data == "admin_add_schedule")
async def admin_add_schedule_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        for uid in user_ids:
            if user_status.get(uid) != 'blocked':
                name = user_names.get(uid, f"ID: {uid}")
                specialty = user_specialty.get(uid, '')
                specialty_display = f" [{specialty}]" if specialty else ""
                builder.row(
                    InlineKeyboardButton(text=f"👤 {name}{specialty_display}", callback_data=f"admin_teacher_{uid}")
                )
        
        if not builder.buttons:
            await callback.message.edit_text("📭 Faol o'qituvchilar yo'q.")
            await callback.answer()
            return
        
        await state.set_state(AdminAddSchedule.selecting_teacher)
        await callback.message.edit_text(
            "👤 O'qituvchini tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_add_schedule_start error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(AdminAddSchedule.selecting_teacher, F.data.startswith("admin_teacher_"))
async def admin_add_schedule_teacher(callback: types.CallbackQuery, state: FSMContext):
    try:
        teacher_id = int(callback.data.replace("admin_teacher_", ""))
        await state.update_data(teacher_id=teacher_id)
        
        builder = InlineKeyboardBuilder()
        for location in LOCATIONS:
            builder.row(
                InlineKeyboardButton(text=location['name'], callback_data=f"admin_branch_{location['name']}")
            )
        
        await state.set_state(AdminAddSchedule.selecting_branch)
        await callback.message.edit_text(
            "🏢 Filialni tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_add_schedule_teacher error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(AdminAddSchedule.selecting_branch, F.data.startswith("admin_branch_"))
async def admin_add_schedule_branch(callback: types.CallbackQuery, state: FSMContext):
    try:
        branch = callback.data.replace("admin_branch_", "")
        await state.update_data(branch=branch)
        
        user_id = callback.from_user.id
        lang = user_languages.get(user_id, 'uz')
        lesson_types = LESSON_TYPES.get(lang, LESSON_TYPES['uz'])
        
        builder = InlineKeyboardBuilder()
        for lesson in lesson_types:
            builder.row(
                InlineKeyboardButton(text=lesson, callback_data=f"admin_lesson_{lesson}")
            )
        
        await state.set_state(AdminAddSchedule.selecting_lesson_type)
        await callback.message.edit_text(
            "📚 Dars turini tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_add_schedule_branch error: {e}")
        await callback.answer("❌ Xatolik yuz berdi")

@dp.callback_query(AdminAddSchedule.selecting_lesson_type, F.data.startswith("admin_lesson_"))
async def admin_add_schedule_lesson(callback: types.CallbackQuery, state: FSMContext):
    try:
        lesson_type = callback.data.replace("admin_lesson_", "")
        await state.update_data(lesson_type=lesson_type)
        
        user_id = callback.from_user.id
        lang = user_languages.get(user_id, 'uz')
        weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
        
        builder = InlineKeyboardBuilder()
        for i, day in enumerate(weekdays):
            builder.row(
                InlineKeyboardButton(text=f"⬜ {day}", callback_data=f"admin_weekday_{i}")
            )
        builder.row(
            InlineKeyboardButton(text="➡️ Keyingisi", callback_data="admin_weekdays_next")
        )
        
        await state.update_data(selected_days={})
        await state.set_state(AdminAddSchedule.selecting_weekdays)
        await callback.message.edit_text(
            "📅 Qaysi kunlarda dars bor?",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_add_schedule_lesson error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(AdminAddSchedule.selecting_weekdays, F.data.startswith("admin_weekday_"))
async def admin_add_schedule_weekday_select(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        selected_days = data.get('selected_days', {})
        day_index = int(callback.data.replace("admin_weekday_", ""))
        
        user_id = callback.from_user.id
        lang = user_languages.get(user_id, 'uz')
        weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
        
        if day_index in selected_days:
            del selected_days[day_index]
        else:
            selected_days[day_index] = None
        
        await state.update_data(selected_days=selected_days)
        
        builder = InlineKeyboardBuilder()
        for i, day in enumerate(weekdays):
            if i in selected_days:
                builder.row(
                    InlineKeyboardButton(text=f"✅ {day}", callback_data=f"admin_weekday_{i}")
                )
            else:
                builder.row(
                    InlineKeyboardButton(text=f"⬜ {day}", callback_data=f"admin_weekday_{i}")
                )
        builder.row(
            InlineKeyboardButton(text="➡️ Keyingisi", callback_data="admin_weekdays_next")
        )
        
        await callback.message.edit_reply_markup(reply_markup=builder.as_markup())
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_add_schedule_weekday_select error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(AdminAddSchedule.selecting_weekdays, F.data == "admin_weekdays_next")
async def admin_add_schedule_weekdays_next(callback: types.CallbackQuery, state: FSMContext):
    try:
        data = await state.get_data()
        selected_days = data.get('selected_days', {})
        
        if not selected_days:
            await callback.answer("Hech bo'lmaganda 1 kun tanlang!", show_alert=True)
            return
        
        days_without_time =[day for day in selected_days if selected_days[day] is None]
        
        if days_without_time:
            await state.update_data(current_day=days_without_time[0])
            await state.set_state(AdminAddSchedule.entering_time)
            
            user_id = callback.from_user.id
            lang = user_languages.get(user_id, 'uz')
            weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
            day_name = weekdays[days_without_time[0]]
            
            await callback.message.edit_text(
                f"⏰ {day_name} kuni soat nechida?\n\nFormat: HH:MM (masalan: 09:00)"
            )
        else:
            await admin_save_new_schedule(callback.message, state)
        
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_add_schedule_weekdays_next error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(AdminAddSchedule.entering_time)
async def admin_add_schedule_enter_time(message: types.Message, state: FSMContext):
    try:
        time_str = message.text.strip()
        hours, minutes = map(int, time_str.split(':'))
        if hours < 0 or hours > 23 or minutes < 0 or minutes > 59:
            raise ValueError
        formatted_time = f"{hours:02d}:{minutes:02d}"
    except:
        await message.answer("❌ Noto'g'ri format! Iltimos, HH:MM formatida kiriting (masalan: 09:00)")
        return
    
    data = await state.get_data()
    selected_days = data.get('selected_days', {})
    current_day = data.get('current_day')
    
    selected_days[current_day] = formatted_time
    await state.update_data(selected_days=selected_days)
    
    days_without_time = [day for day in selected_days if selected_days[day] is None]
    
    if days_without_time:
        await state.update_data(current_day=days_without_time[0])
        
        user_id = message.from_user.id
        lang = user_languages.get(user_id, 'uz')
        weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
        day_name = weekdays[days_without_time[0]]
        
        await message.answer(
            f"⏰ {day_name} kuni soat nechida?\n\nFormat: HH:MM (masalan: 09:00)"
        )
    else:
        await admin_save_new_schedule(message, state)

async def admin_save_new_schedule(message: types.Message, state: FSMContext):
    try:
        data = await state.get_data()
        teacher_id = data.get('teacher_id')
        branch = data.get('branch')
        lesson_type = data.get('lesson_type')
        selected_days = data.get('selected_days', {})
        
        user_id = message.from_user.id
        lang = user_languages.get(user_id, 'uz')
        weekdays = WEEKDAYS.get(lang, WEEKDAYS['uz'])
        
        schedule_id = f"schedule_{teacher_id}_{datetime.now().timestamp()}"
        
        days_with_names = {}
        for day_index, time in selected_days.items():
            day_name = weekdays[day_index]
            days_with_names[day_name] = time
        
        schedules[schedule_id] = {
            'user_id': teacher_id,
            'branch': branch,
            'lesson_type': lesson_type,
            'days': days_with_names
        }
        user_schedules[teacher_id].append(schedule_id)
        
        await db.save_schedule(schedule_id, teacher_id, branch, lesson_type, days_with_names)
        
        try:
            await bot.send_message(
                teacher_id,
                get_text(teacher_id, 'schedule_updated')
            )
            
            pdf_buffer = await create_schedule_pdf(teacher_id)
            clean_name = clean_pdf_text(user_names.get(teacher_id, 'user'))
            await bot.send_document(
                teacher_id,
                types.BufferedInputFile(pdf_buffer.getvalue(), 
                                        filename=f"Dars_Jadvali_{clean_name}.pdf"),
                caption="📅 Yangi dars jadvalingiz (Landscape formatda)"
            )
        except Exception as e:
            logging.error(f"Failed to notify teacher {teacher_id}: {e}")
        
        await message.answer(f"✅ Dars jadvali muvaffaqiyatli qo'shildi!")
        
        await state.clear()
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Admin panel", callback_data="admin_back"))
        await message.answer("Admin panelga qaytish:", reply_markup=builder.as_markup())
    except Exception as e:
        logging.error(f"admin_save_new_schedule error: {e}")
        await message.answer("❌ Jadvalni saqlashda xatolik yuz berdi")
        await state.clear()

@dp.callback_query(F.data == "admin_locations_main")
async def admin_locations_main(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="➕ Yangi filial qo'shish", callback_data="admin_location_add"),
            InlineKeyboardButton(text="📋 Barcha filiallar", callback_data="admin_location_list")
        )
        builder.row(
            InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_back")
        )
        
        await callback.message.edit_text(
            "🏢 Filiallarni boshqarish",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_locations_main error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.callback_query(F.data == "admin_location_list")
async def admin_location_list(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        
        for loc in LOCATIONS:
            maps_link = get_yandex_maps_link(loc['lat'], loc['lon'])
            builder.row(
                InlineKeyboardButton(text=f"📍 {loc['name']}", url=maps_link)
            )
        
        builder.row(InlineKeyboardButton(text="🔙 Ortga", callback_data="admin_locations_main"))
        
        await callback.message.edit_text(
            "📋 Barcha filiallar (lokatsiya uchun bosing):",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_location_list error: {e}")
        await callback.message.edit_text("❌ Filiallar ro'yxatini olishda xatolik yuz berdi")
        await callback.answer()

@dp.callback_query(F.data == "admin_location_add")
async def admin_location_add_start(callback: types.CallbackQuery, state: FSMContext):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        await state.set_state(AddLocation.waiting_for_name)
        await callback.message.edit_text(
            "🏢 Yangi filial nomini kiriting:"
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_location_add_start error: {e}")
        await callback.answer("Xatolik yuz berdi")

@dp.message(AddLocation.waiting_for_name)
async def admin_location_add_name(message: types.Message, state: FSMContext):
    if not check_admin(message.chat.id):
        await state.clear()
        return
    
    try:
        await state.update_data(name=message.text.strip())
        await state.set_state(AddLocation.waiting_for_coords)
        await message.answer(
            "📍 Filial koordinatalarini kiriting (format: lat,lon)\n"
            "Masalan: 41.315790,69.209515"
        )
    except Exception as e:
        logging.error(f"admin_location_add_name error: {e}")
        await message.answer("❌ Xatolik yuz berdi")
        await state.clear()

@dp.message(AddLocation.waiting_for_coords)
async def admin_location_add_coords(message: types.Message, state: FSMContext):
    if not check_admin(message.chat.id):
        await state.clear()
        return
    
    try:
        lat, lon = map(float, message.text.strip().split(','))
        data = await state.get_data()
        name = data['name']
        
        LOCATIONS.append({"name": name, "lat": lat, "lon": lon})
        
        await message.answer(f"✅ Filial muvaffaqiyatli qo'shildi!\n\n{name}\n📍 {lat}, {lon}")
        
        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(text="🔙 Admin panel", callback_data="admin_back"))
        await message.answer("Admin panelga qaytish:", reply_markup=builder.as_markup())
        
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}\nQaytadan urinib ko'ring.")
        return
    
    await state.clear()

@dp.callback_query(F.data == "admin_back")
async def admin_back(callback: types.CallbackQuery):
    if not check_admin(callback.message.chat.id):
        await callback.answer("Ruxsat yo'q!")
        return
    
    try:
        builder = InlineKeyboardBuilder()
        builder.row(
            InlineKeyboardButton(text="💰 Oylik hisoblash", callback_data="admin_salary_calc")
        )
        builder.row(
            InlineKeyboardButton(text="🖼 Vizual Jadval (Haftalik)", callback_data="admin_visual_schedule")
        )
        builder.row(
            InlineKeyboardButton(text="➕ Guruh shakllantirish", callback_data="admin_create_group")
        )
        builder.row(
            InlineKeyboardButton(text="👥 Foydalanuvchilar", callback_data="admin_users_main"),
            InlineKeyboardButton(text="📢 Xabar yuborish", callback_data="admin_broadcast")
        )
        builder.row(
            InlineKeyboardButton(text="🏢 Filiallar", callback_data="admin_locations_main"),
            InlineKeyboardButton(text="📅 Dars jadvallari", callback_data="admin_schedules_main")
        )
        builder.row(
            InlineKeyboardButton(text="📊 Oylik hisobot (Excel)", callback_data="admin_excel_menu"),
            InlineKeyboardButton(text="📊 Kunlik PDF", callback_data="admin_pdf_report")
        )
        
        await callback.message.edit_text(
            "👨‍💼 Admin Panel\n\nKerakli bo'limni tanlang:",
            reply_markup=builder.as_markup()
        )
        await callback.answer()
    except Exception as e:
        logging.error(f"admin_back error: {e}")
        await callback.message.edit_text("❌ Admin panelga qaytishda xatolik yuz berdi")
        await callback.answer()

async def auto_daily_report_task():
    while True:
        now = datetime.now(UZB_TZ)
        if now.hour == 10 and now.minute == 10:
            yesterday = now.date() - timedelta(days=1)
            logging.info(f"Avtomatik hisobot yuborilmoqda: {yesterday}")
            
            try:
                pdf_buf = await get_combined_report_pdf(yesterday)
                await bot.send_document(
                    chat_id=ADMIN_GROUP_ID,
                    document=types.BufferedInputFile(pdf_buf.read(), filename=f"hisobot_{yesterday}.pdf"),
                    caption=f"📅 Kechagi kun ({yesterday}) uchun avtomatik davomat hisoboti."
                )
            except Exception as e:
                logging.error(f"Auto report xatosi: {e}")
            
            await asyncio.sleep(61)
        
        await asyncio.sleep(30)

async def check_schedule_reminders():
    while True:
        now_uzb = datetime.now(UZB_TZ)
        current_time = now_uzb.strftime("%H:%M")
        today_date = now_uzb.strftime("%Y-%m-%d")
        current_day_name = WEEKDAYS_UZ[now_uzb.weekday()]
        
        for schedule_id, schedule in schedules.items():
            user_id = schedule['user_id']
            if user_status.get(user_id) == 'blocked':
                continue
            
            branch = schedule['branch']
            days = schedule['days']
            
            if current_day_name in days:
                lesson_time = days[current_day_name]
                lesson_dt = datetime.strptime(lesson_time, "%H:%M")
                
                remind_dt = lesson_dt - timedelta(minutes=1)
                remind_time = remind_dt.strftime("%H:%M")
                
                check_dt = lesson_dt + timedelta(minutes=1)
                check_time = check_dt.strftime("%H:%M")
                
                if current_time == remind_time:
                    msg = (f"🔔 ESLATMA\n\n"
                           f"Bugun soat {lesson_time} da {branch} filialida darsingiz boshlanmoqda.\n"
                           f"📍 Iltimos, darsingizni davomatini qilishni unutmang!")
                    try:
                        await bot.send_message(user_id, msg)
                        logging.info(f"1-min reminder sent to {user_id} for {branch} at {lesson_time}")
                    except Exception as e:
                        logging.error(f"Failed to send reminder to {user_id}: {e}")
                
                elif current_time == check_time:
                    attended = any(k[0] == user_id and k[1] == branch and k[2] == today_date for k in daily_attendance_log)
                    if not attended:
                        msg = (f"⚠️ DIQQAT: DAVOMAT QILINMADI!\n\n"
                               f"Darsingiz soat {lesson_time} da {branch} filialida boshlangan.\n"
                               f"Hozirgi vaqt: {now_uzb.strftime('%H:%M')}.\n\n"
                               f"Iltimos, darhol davomatni tasdiqlang!")
                        try:
                            await bot.send_message(user_id, msg)
                            logging.info(f"Late reminder sent to {user_id} for {branch}")
                        except Exception as e:
                            logging.error(f"Failed to send late reminder to {user_id}: {e}")
        
        await asyncio.sleep(60)

async def main():
    await db.create_pool()
    await db.init_tables()
    await db.load_to_ram()
    
    asyncio.create_task(start_web_server())
    asyncio.create_task(check_schedule_reminders())
    asyncio.create_task(auto_daily_report_task())
    
    await bot.delete_webhook(drop_pending_updates=True)
    
    await dp.start_polling(bot, skip_updates=True)

if __name__ == "__main__":
    asyncio.run(main())
